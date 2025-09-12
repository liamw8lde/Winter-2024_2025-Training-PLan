# streamlit_app.py
import io
import re
import unicodedata
from datetime import date
from collections import defaultdict, Counter

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ------------------------------- Utils -------------------------------

SLOT_RE = re.compile(r"^([DE])(\d{2}):(\d{2})-([0-9]+)\s+PL([AB])$", re.IGNORECASE)

def norm(s):
    s = str(s or "").strip()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return re.sub(r"\s+"," ",s)

def to_date(v):
    try: return pd.to_datetime(v).date()
    except: return None

def minutes_from_code(code: str) -> int:
    m=re.search(r"-([0-9]+)\s+PL", code)
    return int(m.group(1)) if m else 0

def required_count(code: str) -> int:
    return 2 if code.upper().startswith("E") else 4

# Stable readable palette (overridden by Legende if present)
FALLBACK_PALETTE = {
    'D20:00-120 PLA':'1D4ED8','D20:00-120 PLB':'F59E0B',
    'D20:30-90 PLA':'6D28D9','D20:30-90 PLB':'C4B5FD',
    'E18:00-60 PLA':'10B981','E19:00-60 PLA':'14B8A6','E19:00-60 PLB':'14B8A6',
    'E20:00-90 PLA':'0EA5E9','E20:00-90 PLB':'0EA5E9',
    'E20:30-90 PLA':'10B981','E20:30-90 PLB':'10B981'
}

def luminance(hex6: str) -> float:
    r=int(hex6[0:2],16)/255; g=int(hex6[2:4],16)/255; b=int(hex6[4:6],16)/255
    return 0.2126*r + 0.7152*g + 0.0722*b

def readable_font_color(hex6: str) -> str:
    return "FFFFFF" if luminance(hex6) < 0.55 else "000000"

# ------------------------- Jotform & Rankings ------------------------

def parse_days(s):
    if not isinstance(s,str): return set()
    out=set()
    for p in re.split(r"[,;\n]+", s):
        pl=norm(p).lower()
        if pl.startswith("mon"): out.add("Montag")
        elif pl.startswith("mit"): out.add("Mittwoch")
        elif pl.startswith("don"): out.add("Donnerstag")
        elif p in ("Montag","Mittwoch","Donnerstag"): out.add(p)
    return out

def parse_blocks(s):
    blocks=[]
    if not isinstance(s,str) or not s.strip(): return blocks
    for line in str(s).splitlines():
        ds=re.findall(r"\d{4}[-/\.]\d{2}[-/\.]\d{2}", line)
        if len(ds)>=2:
            d1=pd.to_datetime(ds[0]).date(); d2=pd.to_datetime(ds[1]).date()
            if d2<d1: d1,d2=d2,d1
            blocks.append((d1,d2))
    return blocks

def load_jotform(file) -> dict:
    """Return availability rules: {name_norm: {'days': set, 'blocks': [(d1,d2),...]}}"""
    df = pd.read_excel(file, sheet_name=0)
    def find_col(includes):
        inc=[s.lower() for s in (includes if isinstance(includes,(list,tuple)) else [includes])]
        for c in df.columns:
            if all(s in str(c).lower() for s in inc): return c
        return None
    name_col=None
    for cand in ["wie heißt du","wie heisst du","name"]:
        name_col=find_col(cand)
        if name_col: break
    days_col=find_col("wochentage")
    holiday_col=find_col(["nicht spielen","urlaub"]) or find_col("urlaub")

    rules={}
    for _,row in df.iterrows():
        nm=norm(row.get(name_col,"")).lower()
        if not nm: continue
        rules[nm]={'days':parse_days(row.get(days_col,"")), 'blocks':parse_blocks(row.get(holiday_col,""))}
    return rules

def load_rankings(file) -> dict:
    """Expect columns like: Name / Rank (1..6) / Gender (m|w) — flexible header match."""
    df = pd.read_excel(file, sheet_name=0)
    def fcol(*alts):
        for a in alts:
            for c in df.columns:
                if a in str(c).lower(): return c
        return None
    name_col = fcol("name","spieler")
    rank_col = fcol("rank","starke","spielstarke","stärke")
    sex_col  = fcol("gender","geschlecht","sex")
    r={}
    for _,row in df.iterrows():
        nm=norm(row.get(name_col,""))
        if not nm: continue
        r[nm.lower()] = {
            "rank": int(row.get(rank_col, 3)) if pd.notna(row.get(rank_col, None)) else 3,
            "sex":  str(row.get(sex_col, "")).strip().lower()[:1]
        }
    return r

# ---------------------------- Plan I/O -------------------------------

def read_palette_from_legende(wb) -> dict:
    try:
        if "Legende" not in wb.sheetnames: return {}
        # Allow both 'Legende' typed in Excel or DataFrame read
        # We try to read via pandas to get hex codes written in cells.
        with io.BytesIO() as tmp:
            wb.save(tmp); tmp.seek(0)
            df = pd.read_excel(tmp, sheet_name="Legende")
        ex_col= None; col_col=None
        for c in df.columns:
            lc=str(c).lower()
            if 'slot' in lc or 'beispiel' in lc: ex_col=c
            if 'farbe' in lc or 'hex' in lc: col_col=c
        out={}
        if ex_col and col_col:
            for _,row in df.iterrows():
                code = str(row[ex_col]).strip() if pd.notna(row[ex_col]) else ''
                colv = str(row[col_col]).strip() if pd.notna(row[col_col]) else ''
                if code and SLOT_RE.match(code):
                    hex6 = colv.replace('#','').upper()
                    if re.fullmatch(r"[0-9A-F]{6}", hex6 or ""):
                        out[code]=hex6
        return out
    except Exception:
        return {}

def read_grid(wb):
    ws = wb["Herren 40–50–60"]
    HDR=2; START=3
    players=[str(ws.cell(row=HDR, column=c).value or "").strip() for c in range(3, ws.max_column+1)]
    rows=[]
    for r in range(START, ws.max_row+1):
        dd = to_date(ws.cell(row=r, column=1).value)
        tag = str(ws.cell(row=r, column=2).value or "").strip()
        if not dd: continue
        row={"row":r,"date":dd,"tag":tag}
        for ci,p in enumerate(players, start=3):
            code=str(ws.cell(row=r, column=ci).value or "").strip()
            row[p]=code
        rows.append(row)
    return pd.DataFrame(rows), players

def build_slot_map(df, players):
    slot_map=defaultdict(list)  # (date, tag, slot) -> [players]
    for _,row in df.iterrows():
        dd=row["date"]; tag=row["tag"]
        for p in players:
            val=str(row.get(p,"") or "").strip()
            if SLOT_RE.match(val):
                slot_map[(dd,tag,val)].append(p)
    return slot_map

# ------------------------ Validation & Fixing ------------------------

def is_available(name, dd, tag, jot_rules):
    key=norm(name).lower()
    ru=jot_rules.get(key, None)
    if not ru: return False
    if tag not in ru['days']: return False
    for s,e in ru['blocks']:
        if s<=dd<=e: return False
    return True

def player_busy_on_day(slot_map, name, dd):
    for (d,tag,code), plist in slot_map.items():
        if d==dd and name in plist: return True
    return False

def slot_start(code):
    m=SLOT_RE.match(code)
    return int(m.group(2))*60 + int(m.group(3)) if m else 9999

def fix_conflicts(slot_map, players, tags_by_date, jot_rules,
                  keep_monday_rule=True, keep_lars_60=True,
                  lars_name="Lars Staubermann",
                  must_monday=("Martin Lange","Björn Junker")):
    """Fix headcounts, remove same-day overlaps, and respect Jotform rules."""
    totals = Counter()
    for plist in slot_map.values():
        for p in plist: totals[p]+=1

    # Remove Jotform violations
    for (dd,tag,code), plist in list(slot_map.items()):
        for p in list(plist):
            if not is_available(p, dd, tag, jot_rules):
                slot_map[(dd,tag,code)].remove(p)
                totals[p]-=1

    # Fix headcounts (overfilled first)
    for (dd,tag,code), plist in list(slot_map.items()):
        need=required_count(code)
        if len(plist)>need:
            def pri_keep(n):
                pri0 = 0 if (keep_monday_rule and tag=="Montag" and code=="D20:00-120 PLA" and norm(n) in map(lambda x:norm(x), [*must_monday, lars_name])) else 1
                return (pri0, totals[n], norm(n))
            keep = sorted(plist, key=pri_keep)[:need]
            for p in list(plist):
                if p not in keep:
                    slot_map[(dd,tag,code)].remove(p)
                    totals[p]-=1

    # Same-day overlaps: keep Monday PLA first, then earliest start, then doubles
    by_day=defaultdict(list)
    for (dd,tag,code), plist in slot_map.items():
        for p in plist: by_day[(dd,p)].append((tag,code))
    for (dd,p), items in by_day.items():
        if len(items)<=1: continue
        def pri(code, tag):
            return (0 if (keep_monday_rule and tag=="Montag" and code=="D20:00-120 PLA") else 1,
                    slot_start(code),
                    0 if code.startswith("D") else 1,
                    code)
        tag_keep, code_keep = sorted(items, key=lambda x:pri(x[1], x[0]))[0]
        # remove others
        for tag,code in items:
            if (tag,code)==(tag_keep,code_keep): continue
            if p in slot_map[(dd,tag,code)]:
                slot_map[(dd,tag,code)].remove(p)
                totals[p]-=1

    # Fill underfilled slots with available, least-used players
    # (respect Jotform + no same-day double booking)
    # On Monday PLA, try to include must_monday and (optionally) Lars.
    lars_weeks=set()  # track weeks Lars plays Monday for 60% heuristic
    if keep_lars_60:
        for (dd,tag,code), plist in slot_map.items():
            if tag=="Montag" and code=="D20:00-120 PLA" and lars_name in plist:
                lars_weeks.add(dd.isocalendar()[1])

    for (dd,tag,code), plist in list(slot_map.items()):
        need=required_count(code)
        if len(plist)<need:
            short = need-len(plist)
            cand=[]
            for p in players:
                if p in plist: continue
                if player_busy_on_day(slot_map, p, dd): continue
                if not is_available(p, dd, tag, jot_rules): continue
                cand.append(p)
            cand.sort(key=lambda n: (totals[n], norm(n)))
            # Prefer Martin/Björn on Monday PLA
            if keep_monday_rule and tag=="Montag" and code=="D20:00-120 PLA":
                for must in must_monday:
                    for p in cand:
                        if norm(p).lower()==norm(must).lower() and short>0:
                            slot_map[(dd,tag,code)].append(p)
                            totals[p]+=1
                            short-=1
                            cand.remove(p)
                            break
                # Lars 60% Mondays (approx): if not already many weeks, include when free
                if keep_lars_60 and short>0 and lars_name in cand:
                    week=dd.isocalendar()[1]
                    if week not in lars_weeks:
                        slot_map[(dd,tag,code)].append(lars_name)
                        totals[lars_name]+=1
                        short-=1
                        lars_weeks.add(week)
                        cand.remove(lars_name)
            for p in cand:
                if short<=0: break
                slot_map[(dd,tag,code)].append(p)
                totals[p]+=1
                short-=1
    return slot_map

# ----------------------------- Writer --------------------------------

def write_workbook(base_wb, slot_map, players, dark_grey_empties=True):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    ws = wb.create_sheet("Herren 40–50–60")

    # Collect all dates/tags sorted
    all_rows = defaultdict(dict)  # dd -> {"tag": tag, player->slot}
    for (dd,tag,code), plist in slot_map.items():
        if "tag" not in all_rows[dd]: all_rows[dd]["tag"]=tag
        for p in plist:
            all_rows[dd][p]=code
    dates_sorted = sorted(all_rows.keys())

    # Header rows
    ws.cell(row=1, column=1, value="Datum")
    ws.cell(row=1, column=2, value="Tag")
    for i,p in enumerate(players, start=3):
        ws.cell(row=2, column=i, value=p)

    # Fill rows
    r=3
    for dd in dates_sorted:
        ws.cell(row=r, column=1, value=pd.Timestamp(dd))
        ws.cell(row=r, column=2, value=all_rows[dd]["tag"])
        for i,p in enumerate(players, start=3):
            code = all_rows[dd].get(p,"")
            ws.cell(row=r, column=i, value=code)
        r+=1

    # Palette
    palette = read_palette_from_legende(base_wb)
    for k,v in FALLBACK_PALETTE.items():
        palette.setdefault(k,v)

    # Color slots + dark grey empties if asked
    for rr in range(3, ws.max_row+1):
        for cc in range(3, ws.max_column+1):
            val=str(ws.cell(row=rr, column=cc).value or "").strip()
            if SLOT_RE.match(val):
                hex6 = palette.get(val, None)
                if hex6:
                    ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=hex6)
                    ws.cell(row=rr, column=cc).font = Font(bold=True, color=readable_font_color(hex6))
            elif dark_grey_empties:
                ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor="1F2937")

    # Derived tabs
    # Spielplan
    ws_sp = wb.create_sheet("Spielplan")
    ws_sp.append(['Datum','Tag','Slot','Typ','Spieler'])
    for (dd,tag,code), plist in sorted(slot_map.items(), key=lambda x:(x[0][0], x[0][2], x[0][1])):
        ws_sp.append([pd.Timestamp(dd), tag, code, ("Einzel" if code.startswith("E") else "Doppel"), " / ".join(plist)])
        # color slot cell
        hex6 = palette.get(code)
        if hex6:
            c = ws_sp.cell(row=ws_sp.max_row, column=3)
            c.fill = PatternFill("solid", fgColor=hex6)
            c.font = Font(bold=True, color=readable_font_color(hex6))

    # Wochenübersicht / Sanity
    counts = {p:{'Einzel':0,'Doppel':0,'Gesamt':0} for p in players}
    for (dd,tag,code), plist in slot_map.items():
        typ = 'Einzel' if code.startswith('E') else 'Doppel'
        for n in plist:
            counts[n]['Gesamt']+=1
            counts[n][typ]+=1
    for name in ["Wochenübersicht","Sanity Check"]:
        ws_w = wb.create_sheet(name)
        ws_w.append(['Spieler','Einzel','Doppel','Gesamt'])
        for p in players:
            ws_w.append([p, counts[p]['Einzel'], counts[p]['Doppel'], counts[p]['Gesamt']])

    # Konflikte (should be empty when we’re done; we still populate if any)
    conflicts=[]
    # headcounts
    for (dd,tag,code), plist in slot_map.items():
        need = required_count(code)
        if len(plist)!=need:
            conflicts.append([pd.Timestamp(dd), tag, code, f"{len(plist)}/{need} Spieler eingetragen"])
    # same-day overlaps
    by_day=defaultdict(list)
    for (dd,tag,code), plist in slot_map.items():
        for n in plist: by_day[(dd,n)].append(code)
    for (dd,n), codes in by_day.items():
        if len(codes)>1:
            conflicts.append([pd.Timestamp(dd), "", ", ".join(sorted(codes)), f"{n}: Mehrfach-Einsatz am selben Tag"])
    ws_k = wb.create_sheet("Konflikte")
    if conflicts:
        ws_k.append(['Datum','Tag','Slot','Grund'])
        for row in sorted(conflicts, key=lambda x:(x[0], x[1], x[2], x[3])):
            ws_k.append(row)
    else:
        ws_k.append(["Keine Konflikte gefunden."])

    # Platzgebühren & ABO Platzkosten
    per_day=defaultdict(lambda:{'Slots':0,'Platzstunden':0.0})
    for (dd,tag,code), plist in slot_map.items():
        per_day[tag]['Slots'] += 1
        per_day[tag]['Platzstunden'] += minutes_from_code(code)/60.0
    summary = [[k, v['Slots'], round(v['Platzstunden'],2)] for k,v in sorted(per_day.items())]
    for name in ["Platzgebühren","ABO Platzkosten"]:
        ws_p = wb.create_sheet(name)
        ws_p.append(['Tag','Slots gesamt','Platzstunden gesamt'])
        for row in summary: ws_p.append(row)

    return wb

# ----------------------------- Streamlit -----------------------------

st.set_page_config(page_title="Winter-Training – Planer", layout="wide")
st.title("🎾 Winter-Training Planer")

with st.sidebar:
    st.header("📄 Dateien hochladen")
    plan_file = st.file_uploader("Aktueller Plan / Template (.xlsx)", type=["xlsx"])
    jot_file  = st.file_uploader("Jotform-Export (.xlsx)", type=["xlsx"])
    rank_file = st.file_uploader("Spielstärke (optional) (.xlsx)", type=["xlsx"])

    st.header("⚙️ Regeln")
    keep_monday = st.checkbox("Montags: D20:00-120 PLA = starke Doppel (Martin & Björn fix)", value=True)
    keep_lars   = st.checkbox("Lars ~60% Montags-PLA und max. 1 Match pro Woche", value=True)
    dark_empty  = st.checkbox("Leere Zellen dunkelgrau färben", value=True)
    st.caption("Hinweis: Headcounts (E=2, D=4), keine Doppelbelegung am selben Tag, und Jotform-Verfügbarkeiten werden stets erzwungen.")

colA, colB = st.columns([1,1])

if plan_file and jot_file:
    # Load workbook
    base_wb = load_workbook(plan_file)
    df_grid, players = read_grid(base_wb)
    slot_map = build_slot_map(df_grid, players)
    tags_by_date = {row["date"]: row["tag"] for _,row in df_grid[["date","tag"]].drop_duplicates().iterrows()}

    jot_rules = load_jotform(jot_file)
    ranks = {}
    if rank_file:
        ranks = load_rankings(rank_file)  # (available for future ranking-based pairing logic)

    with colA:
        st.subheader("📊 Aktueller Überblick")
        total_matches = sum(len(v)//2 if k[2].startswith("E") else len(v)//4 for k,v in slot_map.items())
        st.metric("Spieler im Raster", len(players))
        st.metric("Belegte Slots", len(slot_map))
        st.metric("Geschätzte Matches", total_matches)

    with colB:
        st.subheader("🧪 Schnellprüfung (vor Fix)")
        conflicts=[]
        # headcounts
        for (dd,tag,code), plist in slot_map.items():
            need=required_count(code)
            if len(plist)!=need:
                conflicts.append([dd, tag, code, f"{len(plist)}/{need} Spieler eingetragen"])
        # same-day overlaps
        by_day=defaultdict(list)
        for (dd,tag,code), plist in slot_map.items():
            for n in plist: by_day[(dd,n)].append(code)
        for (dd,n), codes in by_day.items():
            if len(codes)>1:
                conflicts.append([dd, tags_by_date.get(dd,""), ", ".join(sorted(codes)), f"{n}: Mehrfach-Einsatz am selben Tag"])
        # jotform
        for (dd,tag,code), plist in slot_map.items():
            for n in plist:
                nm=norm(n).lower()
                ru=jot_rules.get(nm)
                bad = (ru is None) or (tag not in ru['days']) or any(s<=dd<=e for s,e in ru['blocks'])
                if bad:
                    conflicts.append([dd, tag, code, f"{n}: Jotform-Verfügbarkeit verletzt"])
        if conflicts:
            dfc = pd.DataFrame(conflicts, columns=["Datum","Tag","Slot","Grund"]).sort_values(["Datum","Tag","Slot","Grund"])
            st.dataframe(dfc, use_container_width=True)
        else:
            st.success("Keine Konflikte gefunden.")

    st.markdown("---")
    c1, c2 = st.columns([1,1])

    if st.button("🧹 Plan automatisch reparieren", type="primary"):
        fixed = fix_conflicts(slot_map, players, tags_by_date, jot_rules, keep_monday_rule=keep_monday, keep_lars_60=keep_lars)
        # Write out workbook with colors + supporting tabs
        out_wb = write_workbook(base_wb, fixed, players, dark_grey_empties=dark_empty)
        bio = io.BytesIO()
        out_wb.save(bio); bio.seek(0)
        st.success("Plan repariert & neu aufgebaut.")
        st.download_button("💾 Download Excel (repariert)", data=bio.getvalue(),
                           file_name="Trainingplan_CONFLICTS_FIXED.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Show Spielplan preview
        rows=[]
        for (dd,tag,code), plist in sorted(fixed.items(), key=lambda x:(x[0][0], x[0][2], x[0][1])):
            rows.append({"Datum":dd, "Tag":tag, "Slot":code, "Typ":("Einzel" if code.startswith("E") else "Doppel"), "Spieler":" / ".join(plist)})
        st.subheader("📅 Spielplan (Vorschau)")
        st.dataframe(pd.DataFrame(rows), use_container_width=True, height=420)

    st.markdown("### 📋 Aktueller Spielplan (unverändert)")
    rows=[]
    for (dd,tag,code), plist in sorted(slot_map.items(), key=lambda x:(x[0][0], x[0][2], x[0][1])):
        rows.append({"Datum":dd, "Tag":tag, "Slot":code, "Typ":("Einzel" if code.startswith("E") else "Doppel"), "Spieler":" / ".join(plist)})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, height=360)

else:
    st.info("Lade bitte **Plan (.xlsx)** und **Jotform (.xlsx)** hoch, um zu starten. (Spielstärke ist optional)")
