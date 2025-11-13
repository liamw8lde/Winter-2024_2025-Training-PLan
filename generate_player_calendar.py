#!/usr/bin/env python3
"""
Generate a Player Calendar Overview (Spieler-Kalender Übersicht) Excel spreadsheet.
Shows each player's matches with E/D (Singles/Doubles), Court (A/B), and Time.
Color-coded for easy reading.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import re

def load_plan():
    """Load the training plan"""
    df = pd.read_csv("Winterplan_2026.csv", dtype=str)

    # Parse dates
    s = df["Datum"].astype(str).str.strip()
    d1 = pd.to_datetime(s, format="%d.%m.%Y", errors="coerce")
    d2 = pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")
    df["Datum_dt"] = d1.fillna(d2)

    # Sort by date
    df = df.sort_values("Datum_dt")

    return df

def extract_match_info(row):
    """Extract E/D, Court, Time from a match row"""
    # Typ: Einzel or Doppel
    typ = row["Typ"]
    if pd.isna(typ):
        match_type = "?"
    elif typ.lower().startswith("einzel"):
        match_type = "E"
    elif typ.lower().startswith("doppel"):
        match_type = "D"
    else:
        match_type = "?"

    # Extract court and time from Slot (e.g., "E19:00-60 PLA")
    slot = str(row.get("Slot", ""))
    court_match = re.search(r"PL([AB])", slot)
    court = court_match.group(1) if court_match else "?"

    time_match = re.search(r"(\d{2}:\d{2})", slot)
    time = time_match.group(1) if time_match else "?"

    return match_type, court, time

def generate_player_calendar():
    """Generate the player calendar overview Excel file"""
    print("Loading training plan...")
    df = load_plan()

    # Explode players into separate rows
    df["Spieler_list"] = df["Spieler"].str.split(",").apply(
        lambda xs: [x.strip() for x in xs if str(x).strip() and x.strip() != ""]
    )
    df_exp = df.explode("Spieler_list")
    df_exp = df_exp[df_exp["Spieler_list"].notna() & (df_exp["Spieler_list"] != "")]

    # Get all unique players
    all_players = sorted(df_exp["Spieler_list"].unique())

    # Get all unique dates
    all_dates = sorted(df_exp["Datum_dt"].dropna().unique())

    print(f"Found {len(all_players)} players and {len(all_dates)} dates")

    # Create a matrix: dates x players (dates in rows, players in columns)
    calendar_data = []

    for date_dt in all_dates:
        date_str = pd.to_datetime(date_dt).strftime("%d.%m.%Y")
        date_row = {"Datum": date_str}

        # For each player, find if they have a match on this date
        for player in all_players:
            # Find matches for this player on this date
            matches = df_exp[
                (df_exp["Spieler_list"] == player) &
                (df_exp["Datum_dt"] == date_dt)
            ]

            if not matches.empty:
                # Should only be one match per player per date
                match = matches.iloc[0]
                match_type, court, time = extract_match_info(match)
                cell_value = f"{match_type} {time} {court}"
            else:
                cell_value = ""

            date_row[player] = cell_value

        calendar_data.append(date_row)

    # Create DataFrame
    df_calendar = pd.DataFrame(calendar_data)

    print(f"Generated calendar with {len(df_calendar)} players and {len(all_dates)} dates")

    # Create Excel workbook with formatting
    print("Creating formatted Excel file...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Spieler-Kalender 2026"

    # Write headers
    headers = list(df_calendar.columns)
    ws.append(headers)

    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for idx, row in df_calendar.iterrows():
        ws.append(list(row))

    # Color coding and styling
    # E (Einzel/Singles) = Light Blue
    # D (Doppel/Doubles) = Light Green
    einzel_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    doppel_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    # Court A = Bold
    # Court B = Normal
    font_a = Font(bold=True, size=10)
    font_b = Font(bold=False, size=10)

    center_alignment = Alignment(horizontal="center", vertical="center")

    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply styling to data cells
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = thin_border

            # First column (dates) - left align and bold
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True, size=10)
                continue

            # Data cells - color code based on match type
            cell_value = str(cell.value or "")
            if cell_value:
                # Color based on E or D
                if cell_value.startswith("E "):
                    cell.fill = einzel_fill
                elif cell_value.startswith("D "):
                    cell.fill = doppel_fill

                # Font based on Court A or B
                if " A" in cell_value:
                    cell.font = font_a
                elif " B" in cell_value:
                    cell.font = font_b

    # Set column widths
    ws.column_dimensions['A'].width = 15  # Dates
    for col_idx in range(2, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 12  # Player columns

    # Freeze first row and first column
    ws.freeze_panes = "B2"

    # Save the workbook
    output_file = "Spieler_Kalender_2026.xlsx"
    wb.save(output_file)
    print(f"✓ Saved calendar to {output_file}")

    # Also save as CSV for reference
    csv_file = "Spieler_Kalender_2026.csv"
    df_calendar.to_csv(csv_file, index=False)
    print(f"✓ Saved CSV version to {csv_file}")

    return output_file, csv_file

if __name__ == "__main__":
    print("=" * 80)
    print("GENERATING SPIELER-KALENDER ÜBERSICHT 2026")
    print("=" * 80)
    print()

    output_file, csv_file = generate_player_calendar()

    print()
    print("=" * 80)
    print("COMPLETE!")
    print("=" * 80)
    print()
    print("Legend:")
    print("  E = Einzel (Singles)")
    print("  D = Doppel (Doubles)")
    print("  A/B = Court A or B")
    print("  Time = Match start time (HH:MM)")
    print()
    print("Color Coding:")
    print("  Light Blue = Singles (E)")
    print("  Light Green = Doubles (D)")
    print("  Bold = Court A")
    print("  Normal = Court B")
    print()
