import streamlit as st
import pandas as pd
import re
import base64, json, requests, io
from datetime import date, datetime, timedelta

# ==================== PAGE CONFIG ====================
st.set_page_config(
    page_title="Training Plan 2026 Auto-Population",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.title("🎾 Training Plan 2026 Auto-Population")
st.caption("Automatically fill empty training slots for Winter 2026 season (Jan - Apr)")

# ==================== CONFIGURATION ====================
EDIT_PASSWORD = "tennis"
COURT_RATE_PER_HOUR = 17.50

# 2026 SEASON SPECIFIC
PLAN_FILE = "Winterplan_2026.csv"
PREFS_FILE = "Spieler_Preferences_2026.csv"
RANK_FILE = "Player_Ranks_2026.csv"
SEASON_START = date(2026, 1, 5)  # First Monday of January 2026
SEASON_END = date(2026, 4, 26)   # Last date in preferences

# Allowed weekly slots
ALLOWED_SLOTS = {
    "Montag": [
        ("D20:00-120 PLA", "Doppel"),
        ("D20:00-120 PLB", "Doppel"),
    ],
    "Mittwoch": [
        ("E18:00-60 PLA",  "Einzel"),
        ("E19:00-60 PLA",  "Einzel"),
        ("E19:00-60 PLB",  "Einzel"),
        ("D20:00-90 PLA",  "Doppel"),
        ("D20:00-90 PLB",  "Doppel"),
    ],
    "Donnerstag": [
        ("E20:00-90 PLA",  "Einzel"),
        ("E20:00-90 PLB",  "Einzel"),
    ],
}

WEEKDAY_TO_ISO = {"Montag": 1, "Mittwoch": 3, "Donnerstag": 4}
BLACKOUT_MMDD = {(12, 24), (12, 25), (12, 31)}

# Player rankings (1 = strongest, 6 = weakest) - FALLBACK if CSV fails
RANK_FALLBACK = {
    "Andreas Dank": 5, "Anke Ihde": 6, "Arndt Stueber": 4, "Bernd Robioneck": 4, "Bernd Sotzek": 3,
    "Bjoern Junker": 1, "Carsten Gambal": 4, "Dirk Kistner": 5, "Frank Koller": 4, "Frank Petermann": 2,
    "Gunnar Brix": 6, "Heiko Thomsen": 4, "Jan Pappenheim": 5, "Jens Hafner": 6, "Jens Krause": 2,
    "Joerg Peters": 2, "Juergen Hansen": 3, "Kai Schroeder": 4, "Karsten Usinger": 5, "Kerstin Baarck": 6,
    "Lars Staubermann": 2, "Lena Meiss": 6, "Liam Wilde": 2, "Lorenz Kramp": 4, "Manfred Grell": 4,
    "Markus Muench": 5, "Martin Lange": 2, "Martina Schmidt": 6, "Matthias Duddek": 3, "Michael Bock": 6,
    "Michael Rabehl": 6, "Mohamad Albadry": 5, "Oliver Boess": 3, "Patrick Buehrsch": 1, "Peter Plaehn": 2,
    "Ralf Colditz": 4, "Sebastian Braune": 6, "Thomas Bretschneider": 2, "Thomas Grueneberg": 2,
    "Tobias Kahl": 5, "Torsten Bartel": 5, "Wolfgang Aleksik": 6
}

WOMEN_SINGLE_BAN = {"Anke Ihde", "Lena Meiss", "Martina Schmidt", "Kerstin Baarck"}

# Women banned from Monday 20:00 doubles Court A (no women in this time slot)
WOMEN_MONDAY_DOUBLES_A_BAN = {"Anke Ihde", "Lena Meiss", "Martina Schmidt", "Kerstin Baarck"}

# Paired players - must play at the same time (different courts OK)
PAIRED_PLAYERS = [
    # ("Lena Meiss", "Kerstin Baarck"),  # Removed: no longer travel together
]

# Partner preferences (enforced for doubles)
PARTNER_PREFERENCES = {
    "Bjoern Junker": ["Martin Lange"],  # Carpool from Schönkirchen
}

# Monthly match limits (player -> max matches per month)
MONTHLY_LIMITS = {
    "Peter Plaehn": 3,  # "2-3 X im Monat" -> max 3
}

# Season match limits (player -> max matches per season)
SEASON_LIMITS = {
    "Torsten Bartel": 0,  # Not playing this winter anymore
    "Patrick Buehrsch": 2,  # Max 2 matches for entire 2026 season
}

# Season match targets (player -> target number of matches)
# Players below their target get priority boost
SEASON_TARGETS = {
    "Thomas Grueneberg": 11,  # "Würde gerne einmal pro Woche spielen!"
}

# Singles variety constraint - max times same pairing can occur
MAX_SINGLES_REPEATS = 3  # Same two players can only play each other max 3 times in singles

# ==================== DATA LOADING ====================
@st.cache_data(show_spinner=False)
def load_plan_csv(file_path):
    """Load and process the training plan CSV"""
    try:
        df = pd.read_csv(file_path, dtype=str)
        return postprocess_plan(df)
    except Exception as e:
        st.error(f"Error loading plan from {file_path}: {e}")
        return None, None

@st.cache_data(show_spinner=False)
def load_preferences_csv(file_path):
    """Load player preferences CSV"""
    try:
        df = pd.read_csv(file_path, dtype=str)
        return df
    except Exception as e:
        st.warning(f"Could not load preferences from {file_path}: {e}")
        return pd.DataFrame()

@st.cache_data(show_spinner=False)
def load_ranks_csv(file_path):
    """Load player rankings from CSV (1=strongest, 6=weakest)"""
    try:
        df = pd.read_csv(file_path)
        rank_dict = {}
        for _, row in df.iterrows():
            name = str(row.get("Spieler", "")).strip()
            rank = row.get("Rank")
            if name and pd.notna(rank):
                rank_dict[name] = int(rank)
        return rank_dict
    except Exception as e:
        st.warning(f"Could not load ranks from {file_path}: {e}. Using fallback.")
        return RANK_FALLBACK

def postprocess_plan(df):
    """Process plan DataFrame and add computed columns"""
    # Make a copy to avoid SettingWithCopyWarning
    df = df.copy()

    required = ["Datum", "Tag", "Slot", "Typ", "Spieler"]
    for c in required:
        df[c] = df[c].astype(str).str.strip()

    # Parse dates (handle both German and ISO formats)
    s = df["Datum"].astype(str).str.strip()
    d1 = pd.to_datetime(s, format="%d.%m.%Y", errors="coerce")
    d2 = pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")
    df["Datum_dt"] = d1.fillna(d2)

    # ISO calendar
    iso = df["Datum_dt"].dt.isocalendar()
    df["Jahr"] = iso["year"]
    df["Woche"] = iso["week"]

    # Extract slot details
    t = df["Slot"].str.extract(r"^([ED])(\d{2}:\d{2})-([0-9]+)\s+PL([AB])$")
    df["S_Art"]   = t[0].fillna("")
    df["S_Time"]  = t[1].fillna("00:00")
    df["S_Dur"]   = t[2].fillna("0")
    df["S_Court"] = t[3].fillna("")

    # Player list
    df["Spieler_list"] = df["Spieler"].str.split(",").apply(
        lambda xs: [x.strip() for x in xs if str(x).strip()]
    )

    # Exploded view
    df_exp = df.explode("Spieler_list").rename(columns={"Spieler_list": "Spieler_Name"})

    return df, df_exp

# ==================== PREFERENCE FUNCTIONS ====================
def get_available_days(df_prefs):
    """Extract available days for each player"""
    result = {}
    if df_prefs.empty:
        return result

    for _, row in df_prefs.iterrows():
        name = str(row.get("Spieler", "")).strip()
        if not name:
            continue
        days_str = str(row.get("AvailableDays", ""))
        days = set()
        for sep in [",", ";"]:
            if sep in days_str:
                days = {d.strip() for d in days_str.split(sep) if d.strip()}
                break
        else:
            if days_str.strip():
                days = {days_str.strip()}
        result[name] = days
    return result

def get_player_preferences(df_prefs):
    """Extract match type preferences"""
    result = {}
    if df_prefs.empty:
        return result

    for _, row in df_prefs.iterrows():
        name = str(row.get("Spieler", "")).strip()
        if not name:
            continue
        pref = str(row.get("Preference", "keine Präferenz")).strip()
        result[name] = pref
    return result

def parse_blocked_ranges(blocked_str):
    """Parse blocked date ranges from CSV"""
    periods = []
    if not blocked_str or pd.isna(blocked_str) or str(blocked_str).strip() == "":
        return periods

    for chunk in str(blocked_str).split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "→" in chunk:
            try:
                s, e = [x.strip() for x in chunk.split("→", 1)]
                sd = datetime.strptime(s, "%Y-%m-%d").date()
                ed = datetime.strptime(e, "%Y-%m-%d").date()
                periods.append((sd, ed))
            except Exception:
                pass
    return periods

def parse_blocked_days(blocked_str):
    """Parse individual blocked days from CSV"""
    periods = []
    if not blocked_str or pd.isna(blocked_str) or str(blocked_str).strip() == "":
        return periods

    for chunk in str(blocked_str).split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        try:
            d = datetime.strptime(chunk, "%Y-%m-%d").date()
            periods.append((d, d))
        except Exception:
            pass
    return periods

def load_holidays(df_prefs):
    """Load all holidays from preferences CSV"""
    holidays = {}
    if df_prefs.empty:
        return holidays

    for _, row in df_prefs.iterrows():
        name = str(row.get("Spieler", "")).strip()
        if not name:
            continue

        blocked_ranges = parse_blocked_ranges(row.get("BlockedRanges", ""))
        blocked_days = parse_blocked_days(row.get("BlockedDays", ""))

        all_periods = blocked_ranges + blocked_days
        if all_periods:
            holidays[name] = holidays.get(name, []) + all_periods

    return holidays

def is_holiday(name, d, holidays_dict):
    """Check if date is a holiday for player"""
    if (d.month, d.day) in BLACKOUT_MMDD:
        return True
    ranges = holidays_dict.get(name, [])
    return any(start <= d <= end for (start, end) in ranges)

# ==================== VALIDATION FUNCTIONS ====================
def week_of(d):
    iso = pd.Timestamp(d).isocalendar()
    return int(iso.year), int(iso.week)

def count_week(df_plan, name, d):
    y, w = week_of(d)
    wk = df_plan[(df_plan["Jahr"] == y) & (df_plan["Woche"] == w)]
    return int(wk["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True).sum())

def count_day(df_plan, name, d):
    """Count matches for a player on a specific date"""
    day_mask = df_plan["Datum_dt"].dt.date == d
    day_df = df_plan[day_mask]
    return int(day_df["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True).sum())

def count_season(df_plan, name):
    return int(df_plan["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True).sum())

def count_month(df_plan, name, d):
    """Count matches for a player in the same month as date d"""
    year = d.year
    month = d.month
    # Filter by year-month
    month_mask = (df_plan["Datum_dt"].dt.year == year) & (df_plan["Datum_dt"].dt.month == month)
    month_df = df_plan[month_mask]
    return int(month_df["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True).sum())

def count_wed20(df_plan, name):
    mask = (
        (df_plan["Tag"] == "Mittwoch") &
        (df_plan["S_Time"] == "20:00") &
        (df_plan["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True))
    )
    return int(mask.sum())

def count_18_19(df_plan, name):
    mask = (
        (df_plan["S_Time"].isin(["18:00", "19:00"])) &
        (df_plan["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True))
    )
    return int(mask.sum())

def count_singles_pairing(df_plan, name1, name2):
    """Count how many times two players have played singles together"""
    # Filter for singles matches only
    singles_mask = df_plan["Typ"].str.lower().str.startswith("einzel")
    singles_df = df_plan[singles_mask]

    # Count matches where both players are present
    count = 0
    for _, row in singles_df.iterrows():
        players_str = row["Spieler"]
        if pd.notna(players_str):
            # Check if both players are in this match
            has_name1 = bool(re.search(fr"\b{re.escape(name1)}\b", players_str))
            has_name2 = bool(re.search(fr"\b{re.escape(name2)}\b", players_str))
            if has_name1 and has_name2:
                count += 1

    return count

def check_violations(name, tag, s_time, typ, df_after, d, available_days, preferences, holidays):
    """Check all violations for a player assignment"""
    violations = []

    # Time conflict check - player cannot be in two places at once
    same_datetime_mask = (
        (df_after["Datum_dt"].dt.date == d) &
        (df_after["S_Time"] == s_time) &
        (df_after["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True))
    )
    if same_datetime_mask.any():
        existing_slots = df_after[same_datetime_mask]["Slot"].unique().tolist()
        # Only flag if player is in MORE THAN ONE slot at the same datetime
        # (Being in one slot is normal, being in two or more is a conflict)
        if len(existing_slots) > 1:
            violations.append(f"{name}: bereits eingeteilt am {d} um {s_time} ({', '.join(existing_slots)}).")

    # Daily limit check - max 1 match per day
    # Note: df_after includes the current assignment, so count > 1 means player already has another match
    day_count = count_day(df_after, name, d)
    if day_count > 1:
        violations.append(f"{name}: max 1 Spiel/Tag überschritten ({day_count} Spiele am {d}).")

    # Weekly limit check - max 1 match per week
    # Note: df_after includes the current assignment, so count > 1 means player already has another match
    week_count = count_week(df_after, name, d)
    if week_count > 1:
        y, w = week_of(d)
        violations.append(f"{name}: max 1 Spiel/Woche überschritten ({week_count} Spiele in Woche {w}/{y}).")

    # Holiday check
    if is_holiday(name, d, holidays):
        violations.append(f"{name}: Urlaub/Blackout am {d}.")

    # Weekday availability
    days = available_days.get(name)
    if days is not None and tag not in days:
        violations.append(f"{name}: nicht verfügbar an {tag}.")

    # Protected player rules
    if name == "Patrick Buehrsch" and s_time != "18:00":
        violations.append(f"{name}: nur 18:00 erlaubt.")
    if name == "Frank Petermann" and s_time not in {"19:00", "20:00"}:
        violations.append(f"{name}: nur 19:00 oder 20:00 erlaubt.")
    if name == "Matthias Duddek" and s_time not in {"18:00", "19:00"}:
        violations.append(f"{name}: nur 18:00 oder 19:00.")
    if name == "Dirk Kistner":
        if tag not in {"Montag", "Mittwoch", "Donnerstag"}:
            violations.append(f"{name}: nur Mo/Mi/Do.")
        if tag == "Mittwoch" and s_time != "19:00":
            violations.append(f"{name}: am Mittwoch nur 19:00.")
    if name == "Arndt Stueber" and not (tag == "Mittwoch" and s_time == "19:00"):
        violations.append(f"{name}: nur Mittwoch 19:00.")
    if name in {"Thommy Grueneberg", "Thomas Grueneberg"}:
        total_after = count_season(df_after, name)
        wed20_after = count_wed20(df_after, name)
        early_after = count_18_19(df_after, name)
        # Only enforce percentage rules after 5+ matches to avoid chicken-egg problem
        if total_after >= 5:
            if wed20_after / total_after > 0.30:
                violations.append(f"{name}: Anteil Mi 20:00 > 30%.")
            if early_after / total_after < 0.70:
                violations.append(f"{name}: Anteil 18/19 < 70%.")
    if name == "Jens Hafner" and not (tag == "Mittwoch" and s_time == "19:00"):
        violations.append(f"{name}: nur Mittwoch 19:00.")
    if typ.lower().startswith("einzel") and name in WOMEN_SINGLE_BAN:
        violations.append(f"{name}: Frauen dürfen kein Einzel spielen.")
    # Check for Monday 20:00 doubles Court A ban for women
    if tag == "Montag" and s_time == "20:00" and typ.lower().startswith("doppel") and name in WOMEN_MONDAY_DOUBLES_A_BAN:
        # Check if this is Court A (PLA)
        player_slot = df_after[
            (df_after["Datum_dt"].dt.date == d) &
            (df_after["S_Time"] == s_time) &
            (df_after["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True))
        ]
        if not player_slot.empty and "PLA" in player_slot.iloc[-1].get("Slot", ""):
            violations.append(f"{name}: Frauen dürfen nicht am Montag 20:00 Doppel Platz A spielen.")

    # Paired players check - must play at same time (they live together)
    for pair in PAIRED_PLAYERS:
        if name in pair:
            # Find the partner(s)
            partners = [p for p in pair if p != name]
            for partner in partners:
                # Check if partner is scheduled at all on this day
                partner_same_day = df_after[
                    (df_after["Datum_dt"].dt.date == d) &
                    (df_after["Spieler"].str.contains(fr"\b{re.escape(partner)}\b", regex=True))
                ]

                if not partner_same_day.empty:
                    # Partner IS scheduled on this day - must be at same time
                    partner_times = partner_same_day["S_Time"].unique()
                    if s_time not in partner_times:
                        violations.append(f"{name}: muss zur gleichen Zeit wie {partner} spielen (Partner spielt um {', '.join(partner_times)}).")
                else:
                    # Partner is NOT scheduled yet - ensure they CAN be scheduled at same time
                    if not can_schedule_paired_partner(name, d, s_time, df_after, available_days, preferences, holidays):
                        violations.append(f"{name}: Partner {partner} kann nicht zur gleichen Zeit ({s_time}) eingeteilt werden.")

    # Monthly match limits
    if name in MONTHLY_LIMITS:
        max_monthly = MONTHLY_LIMITS[name]
        month_count = count_month(df_after, name, d)
        if month_count >= max_monthly:
            violations.append(f"{name}: max {max_monthly}/Monat überschritten ({month_count} bereits geplant).")

    # Season match limits
    if name in SEASON_LIMITS:
        max_season = SEASON_LIMITS[name]
        season_count = count_season(df_after, name)
        if season_count >= max_season:
            violations.append(f"{name}: max {max_season}/Saison überschritten ({season_count} bereits geplant).")

    # Type preferences
    pref = preferences.get(name, "keine Präferenz")
    if pref == "nur Einzel" and typ.lower().startswith("doppel"):
        violations.append(f"{name}: möchte nur Einzel spielen.")
    elif pref == "nur Doppel" and typ.lower().startswith("einzel"):
        violations.append(f"{name}: möchte nur Doppel spielen.")

    return violations

# ==================== SLOT GENERATION ====================
def generate_allowed_slots_calendar_2026():
    """Generate all allowed slots for 2026 season"""
    out = []
    current_date = SEASON_START

    # Iterate through all dates in the season
    while current_date <= SEASON_END:
        # Get German weekday name
        weekday_iso = current_date.isoweekday()  # 1=Monday, 7=Sunday
        german_weekday = None

        if weekday_iso == 1:
            german_weekday = "Montag"
        elif weekday_iso == 3:
            german_weekday = "Mittwoch"
        elif weekday_iso == 4:
            german_weekday = "Donnerstag"

        # If this is a training day, add all slots for it
        if german_weekday and german_weekday in ALLOWED_SLOTS:
            for slot_code, typ_text in ALLOWED_SLOTS[german_weekday]:
                m = re.search(r"-([0-9]+)\s+PL[AB]$", slot_code)
                minutes = int(m.group(1)) if m else 0
                out.append({
                    "Datum": current_date,
                    "Tag": german_weekday,
                    "Slot": slot_code,
                    "Typ": typ_text,
                    "Minutes": minutes,
                })

        current_date += timedelta(days=1)

    return out

def find_empty_slots(df_plan):
    """Find all empty slots in the plan"""
    allowed = generate_allowed_slots_calendar_2026()
    used_pairs = set(zip(pd.to_datetime(df_plan["Datum_dt"]).dt.date, df_plan["Slot"]))

    empty = []
    for slot in allowed:
        pair = (slot["Datum"], slot["Slot"])
        if pair not in used_pairs:
            empty.append(slot)

    return empty

def can_schedule_paired_partner(name, d, s_time, df_after, available_days, preferences, holidays):
    """Check if a paired player's partner can be scheduled at the same time"""
    for pair in PAIRED_PLAYERS:
        if name in pair:
            partners = [p for p in pair if p != name]
            for partner in partners:
                # Check if partner is already scheduled on this day
                partner_same_day = df_after[
                    (df_after["Datum_dt"].dt.date == d) &
                    (df_after["Spieler"].str.contains(fr"\b{re.escape(partner)}\b", regex=True))
                ]

                if partner_same_day.empty:
                    # Partner not scheduled yet - check if there's an available slot at same time
                    # Look for empty slots at this date/time
                    empty_at_time = [slot for slot in generate_allowed_slots_calendar_2026()
                                    if slot["Datum"] == d]

                    # Filter for same time
                    same_time_slots = []
                    for slot in empty_at_time:
                        time_match = re.search(r"(\d{2}:\d{2})", slot["Slot"])
                        slot_time = time_match.group(1) if time_match else "00:00"
                        if slot_time == s_time:
                            # Check if this slot is already filled
                            is_filled = df_after[
                                (df_after["Datum_dt"].dt.date == d) &
                                (df_after["Slot"] == slot["Slot"])
                            ]
                            if is_filled.empty:
                                same_time_slots.append(slot)

                    if not same_time_slots:
                        # No available slot at the same time for partner
                        return False

                    # Check if partner is available/legal
                    partner_days = available_days.get(partner)
                    if partner_days is not None:
                        tag = same_time_slots[0]["Tag"]
                        if tag not in partner_days:
                            return False

                    if is_holiday(partner, d, holidays):
                        return False

                    partner_week_count = count_week(df_after, partner, d)
                    if partner_week_count >= 1:
                        return False

    return True

# ==================== AUTOPOPULATION ALGORITHM ====================
def select_singles_pair(candidates, df_plan, max_rank_diff=2, max_singles_repeats=3):
    """Select 2 players for singles with rank difference ≤ max_rank_diff and variety constraint"""
    for i, c1 in enumerate(candidates):
        if c1["has_violations"]:
            break
        for c2 in candidates[i+1:]:
            if c2["has_violations"]:
                break
            r1 = c1["rank"]
            r2 = c2["rank"]
            if r1 != 999 and r2 != 999 and abs(r1 - r2) <= max_rank_diff:
                # Check singles variety constraint
                pairing_count = count_singles_pairing(df_plan, c1["name"], c2["name"])
                if pairing_count >= max_singles_repeats:
                    # This pairing has occurred too many times, skip it
                    continue
                return [c1["name"], c2["name"]]
    return None

def select_doubles_team(candidates, num_players=4, max_rank_spread=3):
    """Select 4 players for doubles, enforcing partner preferences and rank difference ≤ max_rank_spread"""
    legal = [c for c in candidates if not c["has_violations"]]
    if len(legal) < num_players:
        return None

    # Check for partner preferences
    selected = []
    remaining = legal.copy()

    # First pass: enforce partner preferences
    for candidate in legal:
        player_name = candidate["name"]
        if player_name in PARTNER_PREFERENCES:
            preferred_partners = PARTNER_PREFERENCES[player_name]

            # Check if any preferred partner is in the legal candidates
            for partner_name in preferred_partners:
                partner_candidate = next((c for c in remaining if c["name"] == partner_name), None)

                if partner_candidate:
                    # Add both the player and their preferred partner
                    if player_name not in [s["name"] for s in selected]:
                        selected.append(candidate)
                    if partner_name not in [s["name"] for s in selected]:
                        selected.append(partner_candidate)

                    # Remove from remaining pool
                    remaining = [c for c in remaining if c["name"] not in [player_name, partner_name]]
                    break

    # Second pass: fill remaining slots from the pool
    # Try to build a team with rank spread ≤ max_rank_spread
    while len(selected) < num_players and remaining:
        candidate = remaining.pop(0)

        # Check rank spread if we add this candidate
        test_selected = selected + [candidate]
        ranks = [c["rank"] for c in test_selected if c["rank"] != 999]

        # If all have valid ranks, check spread
        if len(ranks) == len(test_selected):
            rank_spread = max(ranks) - min(ranks)
            if rank_spread <= max_rank_spread:
                selected.append(candidate)
            # else: skip this candidate, try next one
        else:
            # Some players don't have ranks, accept anyway
            selected.append(candidate)

    if len(selected) >= num_players:
        # Final check: verify rank spread for the team
        final_team = selected[:num_players]
        ranks = [c["rank"] for c in final_team if c["rank"] != 999]
        if len(ranks) == num_players:  # All have valid ranks
            rank_spread = max(ranks) - min(ranks)
            if rank_spread > max_rank_spread:
                return None  # Rank spread too large
        return [c["name"] for c in final_team]

    return None

def select_players_for_slot(df_plan, slot_info, all_players, available_days, preferences, holidays, max_singles_repeats=3):
    """Select appropriate players for a slot. Returns (players, used_extended_rank)"""
    datum = slot_info["Datum"]
    tag = slot_info["Tag"]
    slot_code = slot_info["Slot"]
    typ = slot_info["Typ"]

    time_match = re.search(r"(\d{2}:\d{2})", slot_code)
    slot_time = time_match.group(1) if time_match else "00:00"

    num_players = 4 if typ.lower().startswith("doppel") else 2

    # Score all candidates
    candidates = []
    for name in all_players:
        y, w = week_of(datum)
        wk = df_plan[(df_plan["Jahr"] == y) & (df_plan["Woche"] == w)]
        week_count = int(wk["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True).sum())
        season_count = int(df_plan["Spieler"].str.contains(fr"\b{re.escape(name)}\b", regex=True).sum())
        rk = RANK.get(name, 999)

        # Check if this player's paired partner was just scheduled at same time/date
        paired_boost = 0
        for pair in PAIRED_PLAYERS:
            if name in pair:
                partners = [p for p in pair if p != name]
                for partner in partners:
                    # Check if partner is scheduled at this exact date/time
                    partner_at_same_time = df_plan[
                        (df_plan["Datum_dt"].dt.date == datum) &
                        (df_plan["S_Time"] == slot_time) &
                        (df_plan["Spieler"].str.contains(fr"\b{re.escape(partner)}\b", regex=True))
                    ]
                    if not partner_at_same_time.empty:
                        # Partner is scheduled at same time - give huge priority boost
                        paired_boost = -1000

        # Check if this player has a target and is below it
        target_boost = 0
        if name in SEASON_TARGETS:
            target = SEASON_TARGETS[name]
            if season_count < target:
                # Calculate how far below target (more below = higher boost)
                deficit = target - season_count
                target_boost = -100 * deficit  # Strong boost based on deficit

        # Simulate adding player
        y, w = week_of(datum)
        virtual = pd.DataFrame([{
            "Datum": datum.strftime("%Y-%m-%d"),
            "Datum_dt": pd.Timestamp(datum),
            "Tag": tag,
            "Slot": slot_code,
            "Typ": typ,
            "Spieler": name,
            "S_Time": slot_time,
            "Jahr": y,
            "Woche": w
        }])
        df_virtual = pd.concat([df_plan, virtual], ignore_index=True)

        viol = check_violations(name, tag, slot_time, typ, df_virtual, datum, available_days, preferences, holidays)

        candidates.append({
            "name": name,
            "week": week_count,
            "season": season_count,
            "rank": rk,
            "violations": viol,
            "paired_boost": paired_boost,
            "target_boost": target_boost,
            "has_violations": len(viol) > 0,
        })

    # Filter by type preference
    filtered = []
    for c in candidates:
        pref = preferences.get(c["name"], "keine Präferenz")
        if pref == "nur Einzel" and typ.lower().startswith("doppel"):
            continue
        if pref == "nur Doppel" and typ.lower().startswith("einzel"):
            continue
        filtered.append(c)

    # Sort: legal first, then by paired boost, then by target boost, then by usage (season, week), then rank
    # paired_boost: negative = prefer (partner already scheduled at same time)
    # target_boost: negative = prefer (player below their target match count)
    # Prioritize players with fewer matches for better balance
    filtered.sort(key=lambda x: (x["has_violations"], x["paired_boost"], x["target_boost"], x["season"], x["week"], x["rank"], x["name"]))

    # Select players - first try with normal rank constraints
    if typ.lower().startswith("einzel"):
        # Tier 1: Normal constraints
        players = select_singles_pair(filtered, df_plan, max_rank_diff=2, max_singles_repeats=max_singles_repeats)
        if players is not None:
            return players, False
        # Tier 2: Try with extended repetitions (+1), keep rank difference tight
        players = select_singles_pair(filtered, df_plan, max_rank_diff=2, max_singles_repeats=max_singles_repeats + 1)
        if players is not None:
            return players, True
        # Tier 3: Try with extended rank difference (+1)
        players = select_singles_pair(filtered, df_plan, max_rank_diff=3, max_singles_repeats=max_singles_repeats)
        if players is not None:
            return players, True
        return None, False
    else:
        players = select_doubles_team(filtered, num_players, max_rank_spread=3)
        if players is not None:
            return players, False
        # Try with extended rank spread (+1)
        players = select_doubles_team(filtered, num_players, max_rank_spread=4)
        if players is not None:
            return players, True
        return None, False

def autopopulate_plan(df_plan, max_slots, only_legal, all_players, available_days, preferences, holidays, max_singles_repeats=3):
    """Main autopopulation function"""
    df_result = df_plan.copy()
    empty_slots = find_empty_slots(df_result)

    filled_count = 0
    filled_slots = []
    skipped_slots = []
    extended_rank_slots = []  # Track slots that used extended rank difference

    for slot_info in empty_slots:
        if max_slots and filled_count >= max_slots:
            break

        players, used_extended = select_players_for_slot(
            df_result, slot_info, all_players, available_days, preferences, holidays, max_singles_repeats
        )

        if players is None:
            skipped_slots.append(slot_info)
            continue

        # Check violations if only_legal
        if only_legal:
            time_match = re.search(r"(\d{2}:\d{2})", slot_info["Slot"])
            slot_time = time_match.group(1) if time_match else "00:00"

            has_violations = any(
                check_violations(p, slot_info["Tag"], slot_time, slot_info["Typ"],
                               df_result, slot_info["Datum"], available_days, preferences, holidays)
                for p in players
            )
            if has_violations:
                skipped_slots.append(slot_info)
                continue

        # Add new row
        new_row = pd.DataFrame([{
            "Datum": slot_info["Datum"].strftime("%Y-%m-%d"),
            "Tag": slot_info["Tag"],
            "Slot": slot_info["Slot"],
            "Typ": slot_info["Typ"],
            "Spieler": ", ".join(players),
        }])

        df_result = pd.concat([df_result, new_row], ignore_index=True)
        df_result, _ = postprocess_plan(df_result[["Datum", "Tag", "Slot", "Typ", "Spieler"]])

        filled_count += 1
        filled_slots.append({**slot_info, "players": players})

        # Track if this slot used extended rank difference
        if used_extended:
            extended_rank_slots.append({**slot_info, "players": players})

        # Check if any of the players just scheduled are part of a paired group
        # If yes, try to schedule their partner at the same time
        time_match = re.search(r"(\d{2}:\d{2})", slot_info["Slot"])
        slot_time = time_match.group(1) if time_match else "00:00"

        for player in players:
            for pair in PAIRED_PLAYERS:
                if player in pair:
                    # Find the partner
                    partners = [p for p in pair if p != player]
                    for partner in partners:
                        # Check if partner is already scheduled on this day
                        partner_on_day = df_result[
                            (df_result["Datum_dt"].dt.date == slot_info["Datum"]) &
                            (df_result["Spieler"].str.contains(fr"\b{re.escape(partner)}\b", regex=True))
                        ]

                        if partner_on_day.empty:
                            # Partner not scheduled yet - try to find a slot at the same time
                            for next_slot in empty_slots:
                                if (next_slot["Datum"] == slot_info["Datum"] and
                                    next_slot["Slot"] not in [s["Slot"] for s in filled_slots]):
                                    # Check if this slot is at the same time
                                    next_time_match = re.search(r"(\d{2}:\d{2})", next_slot["Slot"])
                                    next_slot_time = next_time_match.group(1) if next_time_match else "00:00"

                                    if next_slot_time == slot_time:
                                        # Try to schedule the partner in this slot
                                        next_players, next_used_extended = select_players_for_slot(
                                            df_result, next_slot, all_players, available_days, preferences, holidays
                                        )

                                        # Check if partner is in the selected players
                                        if next_players and partner in next_players:
                                            # Verify no violations
                                            if only_legal:
                                                has_violations = any(
                                                    check_violations(p, next_slot["Tag"], next_slot_time, next_slot["Typ"],
                                                                   df_result, next_slot["Datum"], available_days, preferences, holidays)
                                                    for p in next_players
                                                )
                                                if not has_violations:
                                                    # Schedule the partner's slot
                                                    partner_row = pd.DataFrame([{
                                                        "Datum": next_slot["Datum"].strftime("%Y-%m-%d"),
                                                        "Tag": next_slot["Tag"],
                                                        "Slot": next_slot["Slot"],
                                                        "Typ": next_slot["Typ"],
                                                        "Spieler": ", ".join(next_players),
                                                    }])
                                                    df_result = pd.concat([df_result, partner_row], ignore_index=True)
                                                    df_result, _ = postprocess_plan(df_result[["Datum", "Tag", "Slot", "Typ", "Spieler"]])
                                                    filled_count += 1
                                                    filled_slots.append({**next_slot, "players": next_players})

                                                    # Track extended rank for partner slot
                                                    if next_used_extended:
                                                        extended_rank_slots.append({**next_slot, "players": next_players})
                                                    break

    return df_result, filled_slots, skipped_slots, extended_rank_slots

# ==================== GITHUB FUNCTIONS ====================
def github_headers():
    token = st.secrets.get("GITHUB_TOKEN")
    if not token:
        raise RuntimeError("GitHub token missing")
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github+json"}

def github_put_file(csv_bytes, message, file_path):
    """Save CSV to GitHub"""
    repo = st.secrets.get("GITHUB_REPO")
    branch = st.secrets.get("GITHUB_BRANCH", "main")

    if not repo:
        raise RuntimeError("GITHUB_REPO not set in secrets")

    # Get current SHA
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"
    r = requests.get(url, headers=github_headers(), params={"ref": branch}, timeout=20)
    sha = None
    if r.status_code == 200:
        sha = r.json().get("sha")

    # Upload
    payload = {
        "message": message,
        "content": base64.b64encode(csv_bytes).decode("utf-8"),
        "branch": branch,
    }
    if sha:
        payload["sha"] = sha

    r = requests.put(url, headers=github_headers(), data=json.dumps(payload), timeout=20)
    if r.status_code in (200, 201):
        return r.json()
    raise RuntimeError(f"GitHub PUT failed {r.status_code}: {r.text}")

# ==================== MAIN APP ====================
# Password protection
def check_password():
    if st.session_state.get("authenticated", False):
        return True

    with st.form("login_form"):
        st.write("🔒 Diese App ist passwortgeschützt")
        password = st.text_input("Passwort", type="password")
        submitted = st.form_submit_button("Einloggen")

        if submitted:
            if password == EDIT_PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Falsches Passwort")
    return False

if not check_password():
    st.stop()

# Load data
with st.spinner("Lade 2026 Daten..."):
    df_plan, df_exp = load_plan_csv(PLAN_FILE)
    df_prefs = load_preferences_csv(PREFS_FILE)
    RANK = load_ranks_csv(RANK_FILE)

if df_plan is None:
    st.error(f"Konnte {PLAN_FILE} nicht laden!")
    st.info(f"Bitte stelle sicher, dass {PLAN_FILE} im gleichen Ordner ist.")
    st.stop()

# Extract preferences
available_days = get_available_days(df_prefs)
preferences = get_player_preferences(df_prefs)
holidays = load_holidays(df_prefs)
all_players = sorted(df_prefs["Spieler"].dropna().unique().tolist()) if not df_prefs.empty else []

# Show rank source info
if RANK == RANK_FALLBACK:
    st.sidebar.warning(f"⚠️ Rankings from fallback (CSV not found)")
else:
    st.sidebar.success(f"✅ Rankings loaded from {RANK_FILE}")

# Initialize state
if "df_work" not in st.session_state:
    st.session_state.df_work = df_plan.copy()

# ==================== ERROR CHECKING FUNCTION ====================
def check_plan_violations(df_plan, available_days, preferences, holidays):
    """Check all slots in the plan for rule violations"""
    if df_plan.empty or len(df_plan) == 0:
        return []

    violations_list = []

    for idx, row in df_plan.iterrows():
        datum_dt = row.get("Datum_dt")
        if pd.isna(datum_dt):
            continue

        datum = datum_dt.date()
        tag = row.get("Tag", "")
        slot = row.get("Slot", "")
        typ = row.get("Typ", "")
        spieler = row.get("Spieler", "")
        s_time = row.get("S_Time", "00:00")

        # Parse players
        players = [p.strip() for p in str(spieler).split(",") if p.strip()]

        # Check each player
        for player in players:
            viols = check_violations(
                player, tag, s_time, typ, df_plan, datum,
                available_days, preferences, holidays
            )

            if viols:
                for viol in viols:
                    violations_list.append({
                        "Datum": datum,
                        "Tag": tag,
                        "Slot": slot,
                        "Typ": typ,
                        "Spieler": player,
                        "Violation": viol
                    })

        # Check singles rank rule
        if typ.lower().startswith("einzel") and len(players) == 2:
            r1 = RANK.get(players[0], 999)
            r2 = RANK.get(players[1], 999)
            if r1 != 999 and r2 != 999 and abs(r1 - r2) > 2:
                violations_list.append({
                    "Datum": datum,
                    "Tag": tag,
                    "Slot": slot,
                    "Typ": typ,
                    "Spieler": f"{players[0]} vs {players[1]}",
                    "Violation": f"Rang-Differenz zu groß: |{r1} - {r2}| = {abs(r1 - r2)} > 2"
                })

        # Check doubles rank rule
        if typ.lower().startswith("doppel") and len(players) == 4:
            ranks = [RANK.get(p, 999) for p in players]
            valid_ranks = [r for r in ranks if r != 999]
            if len(valid_ranks) == 4:  # All players have ranks
                max_rank = max(valid_ranks)
                min_rank = min(valid_ranks)
                rank_spread = max_rank - min_rank
                if rank_spread > 3:
                    violations_list.append({
                        "Datum": datum,
                        "Tag": tag,
                        "Slot": slot,
                        "Typ": typ,
                        "Spieler": ", ".join(players),
                        "Violation": f"Doppel Rang-Differenz zu groß: {max_rank} - {min_rank} = {rank_spread} > 3"
                    })

    return violations_list

# ==================== HELPER FUNCTIONS FOR CALENDAR VIEW ====================
def parse_blocked_ranges(blocked_str):
    """Parse blocked date ranges from string like '2026-01-01→2026-01-04;2026-02-01→2026-02-05'"""
    if pd.isna(blocked_str) or not blocked_str or str(blocked_str).strip() == "":
        return []

    ranges = []
    parts = str(blocked_str).split(";")
    for part in parts:
        part = part.strip()
        if "→" in part:
            try:
                start_str, end_str = part.split("→")
                start_date = datetime.strptime(start_str.strip(), "%Y-%m-%d").date()
                end_date = datetime.strptime(end_str.strip(), "%Y-%m-%d").date()
                ranges.append((start_date, end_date))
            except:
                continue
    return ranges

def parse_blocked_days(blocked_str):
    """Parse specific blocked days from string like '2026-01-15;2026-01-22'"""
    if pd.isna(blocked_str) or not blocked_str or str(blocked_str).strip() == "":
        return []

    days = []
    parts = str(blocked_str).split(";")
    for part in parts:
        part = part.strip()
        try:
            day = datetime.strptime(part, "%Y-%m-%d").date()
            days.append(day)
        except:
            continue
    return days

def parse_available_days_cal(days_str):
    """Parse available days from string like 'Montag;Mittwoch;Donnerstag'"""
    if pd.isna(days_str) or not days_str or str(days_str).strip() == "":
        return set()

    # Map German day names to English
    day_map = {
        "Montag": "Monday",
        "Dienstag": "Tuesday",
        "Mittwoch": "Wednesday",
        "Donnerstag": "Thursday",
        "Freitag": "Friday",
        "Samstag": "Saturday",
        "Sonntag": "Sunday"
    }

    days = set()
    parts = str(days_str).split(";")
    for part in parts:
        part = part.strip()
        if part in day_map:
            days.add(day_map[part])
    return days

def is_player_blocked_cal(player_name, check_date, prefs_df):
    """Check if a player is blocked (holiday/unavailable) on a specific date"""
    player_prefs = prefs_df[prefs_df["Spieler"] == player_name]

    if player_prefs.empty:
        return False, None

    row = player_prefs.iloc[0]

    # Check blocked ranges
    blocked_ranges = parse_blocked_ranges(row.get("BlockedRanges", ""))
    for start, end in blocked_ranges:
        if start <= check_date <= end:
            return True, "🚫"  # Holiday/Blocked

    # Check specific blocked days
    blocked_days = parse_blocked_days(row.get("BlockedDays", ""))
    if check_date in blocked_days:
        return True, "🚫"  # Holiday/Blocked

    # Check available days (day of week)
    available_days = parse_available_days_cal(row.get("AvailableDays", ""))
    if available_days:  # If available days are specified
        day_name = check_date.strftime("%A")  # Get English day name
        if day_name not in available_days:
            return True, "—"  # Not available this day of week

    return False, None

# ==================== SPREADSHEET VIEW FUNCTION ====================
def create_player_calendar_view(df_plan):
    """Create a pivot table with dates as rows and players as columns
    Format: E/D Time Court (e.g., 'E 19:00 A' or 'D 20:00 B')
    Also shows blocked dates (🚫) and unavailable days (—)
    """
    if df_plan.empty or len(df_plan) == 0:
        return pd.DataFrame()

    # Load preferences for blocked dates
    try:
        prefs_df = load_preferences_csv(PREFS_FILE)
    except:
        prefs_df = pd.DataFrame()

    # Explode to get one row per player
    _, df_exp = postprocess_plan(df_plan[["Datum", "Tag", "Slot", "Typ", "Spieler"]])

    if df_exp.empty:
        return pd.DataFrame()

    # Get all unique dates and players from the plan
    all_dates = sorted(df_exp["Datum_dt"].dropna().unique())
    all_players = sorted(df_exp["Spieler_Name"].dropna().unique())

    # Create display value: E/D Time Court
    def format_match(row):
        # Determine E (Einzel) or D (Doppel)
        typ = row['Typ']
        if pd.isna(typ):
            match_type = "?"
        elif typ.lower().startswith("einzel"):
            match_type = "E"
        elif typ.lower().startswith("doppel"):
            match_type = "D"
        else:
            match_type = "?"

        # Get time and court
        time = row['S_Time'] if row['S_Time'] else "?"
        court = row['S_Court'] if row['S_Court'] else "?"

        return f"{match_type} {time} {court}"

    df_exp["Match_Info"] = df_exp.apply(format_match, axis=1)

    # Convert date to string for display (just date, no day name)
    df_exp["Datum_str"] = df_exp["Datum_dt"].dt.strftime("%d.%m.%Y")

    # Create pivot table: Dates (rows) x Players (columns)
    pivot = df_exp.pivot_table(
        index="Datum_str",
        columns="Spieler_Name",
        values="Match_Info",
        aggfunc=lambda x: " | ".join(x) if len(x) > 1 else x.iloc[0],
        fill_value=""
    )

    # Add blocked/unavailable markers for empty cells
    if not prefs_df.empty:
        date_mapping = dict(zip(df_exp["Datum_str"], df_exp["Datum_dt"]))

        for date_str in pivot.index:
            check_date = pd.to_datetime(date_mapping.get(date_str)).date() if date_str in date_mapping else None
            if check_date:
                for player in pivot.columns:
                    # Only check empty cells (no match scheduled)
                    if pivot.loc[date_str, player] == "":
                        is_blocked, marker = is_player_blocked_cal(player, check_date, prefs_df)
                        if is_blocked and marker:
                            pivot.loc[date_str, player] = marker

    # Sort rows by date
    date_mapping = dict(zip(df_exp["Datum_str"], df_exp["Datum_dt"]))
    sorted_rows = sorted(pivot.index, key=lambda x: date_mapping.get(x, pd.Timestamp.min))
    pivot = pivot.loc[sorted_rows]

    # Sort columns (player names) alphabetically
    pivot = pivot[sorted(pivot.columns)]

    return pivot

# ==================== UI ====================
st.info(f"📅 **Saison:** {SEASON_START.strftime('%d.%m.%Y')} bis {SEASON_END.strftime('%d.%m.%Y')} (Januar - April 2026)")
st.markdown("---")

# Current status
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("Aktueller Plan", f"{len(st.session_state.df_work)} Slots")
with col2:
    st.metric("Spieler verfügbar", len(all_players))
with col3:
    empty_count = len(find_empty_slots(st.session_state.df_work))
    st.metric("Leere Slots", empty_count)
with col4:
    total_possible = len(generate_allowed_slots_calendar_2026())
    st.metric("Gesamt möglich", total_possible)

# Action buttons
if len(st.session_state.df_work) > 0:
    col_clear, col_check, col_spacer = st.columns([1, 1, 2])
    with col_clear:
        if st.button("🗑️ Plan leeren", width='stretch', type="secondary"):
            # Store confirmation state
            st.session_state.show_clear_confirm = True

    with col_check:
        if st.button("⚠️ Regeln prüfen", width='stretch', type="secondary"):
            st.session_state.run_error_check = True

    # Confirmation dialog for clear
    if st.session_state.get("show_clear_confirm", False):
        st.warning("⚠️ **Warnung:** Dies löscht alle {0} Slots aus dem aktuellen Plan!".format(len(st.session_state.df_work)))
        col_yes, col_no, col_space = st.columns([1, 1, 2])
        with col_yes:
            if st.button("✅ Ja, leeren", width='stretch', type="primary"):
                # Create empty dataframe with correct structure
                empty_plan = pd.DataFrame(columns=["Datum", "Tag", "Slot", "Typ", "Spieler"])
                st.session_state.df_work, _ = postprocess_plan(empty_plan)
                st.session_state.show_clear_confirm = False
                st.success("✅ Plan geleert! Bereit für Autopopulation.")
                st.rerun()
        with col_no:
            if st.button("❌ Abbrechen", width='stretch'):
                st.session_state.show_clear_confirm = False
                st.rerun()

# Error checking results
if st.session_state.get("run_error_check", False):
    with st.spinner("Prüfe Regeln..."):
        violations = check_plan_violations(
            st.session_state.df_work,
            available_days,
            preferences,
            holidays
        )

    if not violations:
        st.success("✅ Keine Regelverstöße gefunden! Der Plan ist vollständig regelkonform.")
    else:
        st.error(f"⚠️ **{len(violations)} Regelverstöße gefunden!**")

        # Group by type of violation
        df_violations = pd.DataFrame(violations)

        with st.expander(f"📋 Details zu allen {len(violations)} Verstößen", expanded=True):
            # Show as table
            display_df = df_violations.copy()
            display_df["Datum"] = display_df["Datum"].astype(str)
            st.dataframe(
                display_df[["Datum", "Tag", "Slot", "Spieler", "Violation"]],
                width='stretch',
                height=400
            )

        # Summary by violation type
        with st.expander("📊 Zusammenfassung nach Verstoßtyp"):
            violation_counts = df_violations["Violation"].value_counts().reset_index()
            violation_counts.columns = ["Verstoß", "Anzahl"]
            st.dataframe(violation_counts, width='stretch')

        # Summary by player
        with st.expander("👥 Zusammenfassung nach Spieler"):
            player_counts = df_violations["Spieler"].value_counts().reset_index()
            player_counts.columns = ["Spieler", "Anzahl Verstöße"]
            st.dataframe(player_counts, width='stretch')

    # Close button
    if st.button("❌ Prüfung schließen", width='content'):
        st.session_state.run_error_check = False
        st.rerun()

st.markdown("---")

# Create tabs
tab_auto, tab_calendar, tab_rankings, tab_player = st.tabs(["🤖 Auto-Population", "📅 Spieler-Kalender", "🏆 Spieler-Rankings", "👤 Spieler-Profil"])

with tab_calendar:
    st.header("📅 Spieler-Kalender Übersicht")
    st.caption("Zeigt wann welcher Spieler auf welchem Platz spielt")

    # Legend
    st.info("**Format:** E/D Time Court  |  **E** = Einzel (Singles), **D** = Doppel (Doubles)  |  **A/B** = Court  |  **Time** = HH:MM  |  **🚫** = Holiday/Blocked  |  **—** = Day not available")

    # Use working plan or result plan if available
    display_df = st.session_state.get("df_result", st.session_state.df_work)

    if len(display_df) == 0:
        st.info("📭 Der Plan ist noch leer. Verwende die Auto-Population, um Slots zu füllen.")
    else:
        pivot_df = create_player_calendar_view(display_df)

        if not pivot_df.empty:
            st.dataframe(
                pivot_df,
                width='stretch',
                height=600
            )

            # Download options
            col1, col2 = st.columns(2)

            with col1:
                # CSV download
                csv_bytes = pivot_df.to_csv().encode("utf-8")
                st.download_button(
                    "📥 Als CSV herunterladen",
                    data=csv_bytes,
                    file_name="Spieler_Kalender_2026.csv",
                    mime="text/csv",
                    width='stretch'
                )

            with col2:
                # Excel download with color formatting
                if st.button("📊 Als Excel (mit Farben) herunterladen", width='stretch'):
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        import io

                        # Create workbook
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Spieler-Kalender 2026"

                        # Write data
                        for r in dataframe_to_rows(pivot_df, index=True, header=True):
                            ws.append(r)

                        # Style header row
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF", size=11)
                        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment

                        # Color coding
                        einzel_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
                        doppel_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
                        blocked_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
                        unavailable_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light gray

                        font_a = Font(bold=True, size=10)
                        font_b = Font(bold=False, size=10)
                        font_blocked = Font(bold=False, size=10, color="CC0000")  # Red text
                        font_unavailable = Font(bold=False, size=10, color="808080")  # Gray text

                        center_alignment = Alignment(horizontal="center", vertical="center")
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                        # Apply styling to data cells
                        for row_idx in range(2, ws.max_row + 1):
                            for col_idx in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.alignment = center_alignment
                                cell.border = thin_border

                                # First column (dates) - left align and bold
                                if col_idx == 1:
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                                    cell.font = Font(bold=True, size=10)
                                    continue

                                # Data cells - color code based on content
                                cell_value = str(cell.value or "")
                                if cell_value:
                                    # Blocked/Holiday marker
                                    if cell_value == "🚫":
                                        cell.fill = blocked_fill
                                        cell.font = font_blocked
                                    # Unavailable day marker
                                    elif cell_value == "—":
                                        cell.fill = unavailable_fill
                                        cell.font = font_unavailable
                                    # Singles match
                                    elif cell_value.startswith("E "):
                                        cell.fill = einzel_fill
                                        if " A" in cell_value:
                                            cell.font = font_a
                                        elif " B" in cell_value:
                                            cell.font = font_b
                                    # Doubles match
                                    elif cell_value.startswith("D "):
                                        cell.fill = doppel_fill
                                        if " A" in cell_value:
                                            cell.font = font_a
                                        elif " B" in cell_value:
                                            cell.font = font_b

                        # Set column widths
                        ws.column_dimensions['A'].width = 15  # Dates
                        for col_idx in range(2, ws.max_column + 1):
                            col_letter = ws.cell(row=1, column=col_idx).column_letter
                            ws.column_dimensions[col_letter].width = 12

                        # Freeze panes
                        ws.freeze_panes = "B2"

                        # Save to bytes
                        excel_buffer = io.BytesIO()
                        wb.save(excel_buffer)
                        excel_bytes = excel_buffer.getvalue()

                        st.download_button(
                            "💾 Excel-Datei speichern",
                            data=excel_bytes,
                            file_name="Spieler_Kalender_2026.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width='stretch'
                        )

                        st.success("✓ Excel-Datei generiert! Klicke auf den Button oben zum Herunterladen.")
                        st.caption("🔵 Blau = Einzel (E)  |  🟢 Grün = Doppel (D)  |  **Fett** = Platz A  |  Normal = Platz B")

                    except ImportError:
                        st.error("❌ openpyxl nicht installiert. Verwende CSV-Download stattdessen.")
                    except Exception as e:
                        st.error(f"❌ Fehler beim Erstellen der Excel-Datei: {e}")
        else:
            st.warning("⚠️ Konnte keine Kalenderansicht erstellen.")

with tab_auto:
    # Find empty slots
    empty_slots = find_empty_slots(st.session_state.df_work)

    if empty_slots:
        st.header("📋 Leere Slots für 2026")
        st.write(f"Es gibt **{len(empty_slots)}** leere Slots, die gefüllt werden können.")

        # Show sample
        with st.expander(f"Zeige erste {min(20, len(empty_slots))} leere Slots"):
            for i, slot in enumerate(empty_slots[:20]):
                st.write(f"{i+1}. {slot['Datum']} ({slot['Tag']}) — {slot['Slot']} — {slot['Typ']}")

        st.markdown("---")

        # Settings
        st.header("⚙️ Auto-Population Einstellungen")

        col1, col2 = st.columns(2)
        with col1:
            max_slots = st.number_input(
                "Maximale Anzahl Slots zu füllen",
                min_value=1,
                max_value=len(empty_slots),
                value=min(50, len(empty_slots)),
                help="Wie viele Slots sollen automatisch gefüllt werden?"
            )
        with col2:
            only_legal = st.checkbox(
                "Nur legale Zuweisungen (keine Verstöße)",
                value=True,
                help="Wenn aktiv, werden nur Slots mit 100% regelkonformen Spielern gefüllt"
            )

        # Additional settings
        max_singles_repeats = st.number_input(
            "Maximale Wiederholungen Einzel-Paarungen",
            min_value=1,
            max_value=10,
            value=3,
            help="Wie oft dürfen dieselben zwei Spieler im Einzel gegeneinander antreten?"
        )

        st.markdown("---")

        # Actions
        col_preview, col_reset = st.columns(2)
        with col_preview:
            if st.button("🔍 Vorschau generieren", width='stretch', type="primary"):
                with st.spinner("Generiere Auto-Population für 2026..."):
                    df_result, filled, skipped, extended_rank = autopopulate_plan(
                        st.session_state.df_work,
                        max_slots,
                        only_legal,
                        all_players,
                        available_days,
                        preferences,
                        holidays,
                        max_singles_repeats
                    )
                    st.session_state.df_result = df_result
                    st.session_state.filled_slots = filled
                    st.session_state.skipped_slots = skipped
                    st.session_state.extended_rank_slots = extended_rank
                    st.rerun()

        with col_reset:
            if st.button("🔄 Plan zurücksetzen", width='stretch'):
                st.session_state.df_work = df_plan.copy()
                st.session_state.pop("df_result", None)
                st.session_state.pop("filled_slots", None)
                st.session_state.pop("skipped_slots", None)
                st.session_state.pop("extended_rank_slots", None)
                st.rerun()

        # Show results if available
        if "df_result" in st.session_state:
            st.markdown("---")
            st.header("✅ Ergebnisse für 2026")

            filled = st.session_state.get("filled_slots", [])
            skipped = st.session_state.get("skipped_slots", [])
            extended_rank = st.session_state.get("extended_rank_slots", [])

            col1, col2, col3 = st.columns(3)
            with col1:
                st.success(f"**{len(filled)}** Slots erfolgreich gefüllt")
            with col2:
                if skipped:
                    st.warning(f"**{len(skipped)}** Slots übersprungen")
            with col3:
                if extended_rank:
                    st.info(f"**{len(extended_rank)}** mit erweiterter Rang-Differenz")

            # Filled slots details
            if filled:
                with st.expander(f"✅ Gefüllte Slots ({len(filled)})"):
                    for slot in filled:
                        players = slot["players"]
                        typ = slot["Typ"]

                        # Get player rankings
                        player_ranks = [(p, RANK.get(p, "?")) for p in players]
                        players_with_ranks = [f"{p} (Rang {r})" for p, r in player_ranks]
                        players_str = ", ".join(players_with_ranks)

                        # Calculate rank difference/spread
                        ranks = [r for p, r in player_ranks if r != "?"]
                        rank_info = ""
                        if len(ranks) == len(players):  # All have valid ranks
                            if typ.lower().startswith("einzel") and len(ranks) == 2:
                                diff = abs(ranks[0] - ranks[1])
                                emoji = "✅" if diff <= 2 else "❌"
                                rank_info = f"  📊 Rang-Differenz: {diff} {emoji}"
                            elif typ.lower().startswith("doppel") and len(ranks) == 4:
                                spread = max(ranks) - min(ranks)
                                emoji = "✅" if spread <= 3 else "❌"
                                rank_info = f"  📊 Rang-Spread: {min(ranks)}-{max(ranks)} (Diff: {spread}) {emoji}"

                        st.write(f"**{slot['Datum']}** ({slot['Tag']}) — {slot['Slot']} — {slot['Typ']}")
                        st.write(f"  → {players_str}")
                        if rank_info:
                            st.write(rank_info)
                        st.write("")

            # Skipped slots details
            if skipped:
                with st.expander(f"⚠️ Übersprungene Slots ({len(skipped)})"):
                    for slot in skipped:
                        st.write(f"• {slot['Datum']} ({slot['Tag']}) — {slot['Slot']} — {slot['Typ']}")

            # Extended rank difference slots
            if extended_rank:
                with st.expander(f"📏 Slots mit erweiterter Rang-Differenz ({len(extended_rank)})", expanded=True):
                    st.info("ℹ️ Diese Slots konnten nur durch Erhöhung der Rang-Differenz um +1 gefüllt werden (Einzel: Rang-Diff ≤3 statt ≤2, Doppel: Rang-Spread ≤4 statt ≤3)")
                    for slot in extended_rank:
                        players = slot["players"]
                        typ = slot["Typ"]

                        # Get player rankings
                        player_ranks = [(p, RANK.get(p, "?")) for p in players]
                        players_with_ranks = [f"{p} (Rang {r})" for p, r in player_ranks]
                        players_str = ", ".join(players_with_ranks)

                        # Calculate rank difference/spread
                        ranks = [r for p, r in player_ranks if r != "?"]
                        rank_info = ""
                        if len(ranks) == len(players):  # All have valid ranks
                            if typ.lower().startswith("einzel") and len(ranks) == 2:
                                diff = abs(ranks[0] - ranks[1])
                                emoji = "⚠️" if diff == 3 else "❌"
                                rank_info = f"  📊 Rang-Differenz: {diff} {emoji}"
                            elif typ.lower().startswith("doppel") and len(ranks) == 4:
                                spread = max(ranks) - min(ranks)
                                emoji = "⚠️" if spread == 4 else "❌"
                                rank_info = f"  📊 Rang-Spread: {min(ranks)}-{max(ranks)} (Diff: {spread}) {emoji}"

                        st.write(f"**{slot['Datum']}** ({slot['Tag']}) — {slot['Slot']} — {slot['Typ']}")
                        st.write(f"  → {players_str}")
                        if rank_info:
                            st.write(rank_info)
                        st.write("")

            # Statistics
            st.markdown("---")
            st.subheader("📊 Statistiken")

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Vorher", len(st.session_state.df_work))
            with col2:
                st.metric("Nachher", len(st.session_state.df_result))
            with col3:
                st.metric("Hinzugefügt", len(filled), delta=f"+{len(filled)}")

            # Player distribution
            st.subheader("👥 Spieler-Verteilung 2026")
            _, df_result_exp = postprocess_plan(st.session_state.df_result[["Datum", "Tag", "Slot", "Typ", "Spieler"]])
            player_counts = df_result_exp["Spieler_Name"].value_counts().reset_index()
            player_counts.columns = ["Spieler", "Anzahl Matches"]
            # Add rankings (convert to string to avoid mixed types)
            player_counts["Rang"] = player_counts["Spieler"].map(lambda x: str(RANK.get(x, "?")))
            # Reorder columns
            player_counts = player_counts[["Spieler", "Rang", "Anzahl Matches"]]
            st.dataframe(player_counts, width='stretch', height=400)

            # Save buttons
            st.markdown("---")
            col_save, col_download, col_discard = st.columns(3)

            with col_save:
                if st.button("💾 Auf GitHub speichern (2026)", width='stretch', type="primary"):
                    try:
                        df_to_save = st.session_state.df_result[["Datum", "Tag", "Slot", "Typ", "Spieler"]]
                        csv_bytes = df_to_save.to_csv(index=False).encode("utf-8")
                        msg = f"Auto-populate {len(filled)} slots for 2026 season\n\n🤖 Generated with Claude Code"
                        github_put_file(csv_bytes, msg, PLAN_FILE)
                        st.success(f"✅ Erfolgreich auf GitHub als {PLAN_FILE} gespeichert!")
                        st.balloons()
                        # Clear state
                        st.session_state.df_work = st.session_state.df_result.copy()
                        st.session_state.pop("df_result", None)
                        st.session_state.pop("filled_slots", None)
                        st.session_state.pop("skipped_slots", None)
                    except Exception as e:
                        st.error(f"Fehler beim Speichern: {e}")

            with col_download:
                if "df_result" in st.session_state:
                    df_to_save = st.session_state.df_result[["Datum", "Tag", "Slot", "Typ", "Spieler"]]
                    csv_bytes = df_to_save.to_csv(index=False).encode("utf-8")
                    st.download_button(
                        "📥 CSV herunterladen",
                        data=csv_bytes,
                        file_name="Winterplan_2026_autopopulated.csv",
                        mime="text/csv",
                        width='stretch'
                    )

            with col_discard:
                if st.button("🗑️ Vorschau verwerfen", width='stretch'):
                    st.session_state.pop("df_result", None)
                    st.session_state.pop("filled_slots", None)
                    st.session_state.pop("skipped_slots", None)
                    st.session_state.pop("extended_rank_slots", None)
                    st.rerun()

    else:
        st.success("🎉 Keine leeren Slots für 2026 gefunden! Der Plan ist vollständig.")

with tab_rankings:
    st.header("🏆 Spieler-Rankings Verwaltung")
    st.caption("Verwalte Spieler-Rankings (1=stärkster, 6=schwächster)")

    # Check for missing players
    all_players_from_prefs = set(df_prefs["Spieler"].dropna().unique()) if not df_prefs.empty else set()
    players_with_ranks = set(RANK.keys())
    missing_players = all_players_from_prefs - players_with_ranks

    if missing_players:
        st.warning(f"⚠️ **{len(missing_players)} Spieler ohne Rang gefunden!**")
        with st.expander(f"📋 Spieler ohne Rang ({len(missing_players)})", expanded=True):
            for player in sorted(missing_players):
                st.write(f"• {player}")
            st.info("💡 Füge diese Spieler unten zur Rangliste hinzu.")

    # Initialize session state for rankings if not exists
    if "edited_ranks" not in st.session_state:
        st.session_state.edited_ranks = RANK.copy()

    st.markdown("---")
    st.subheader("📊 Aktuelle Rankings")

    # Create editable dataframe
    rank_data = []
    for player, rank in sorted(st.session_state.edited_ranks.items(), key=lambda x: (x[1], x[0])):
        rank_data.append({"Spieler": player, "Rang": rank})

    rank_df = pd.DataFrame(rank_data)

    # Use data_editor for editing
    st.write("**Bearbeite Rankings:** (Doppelklick auf Rang-Zelle zum Ändern)")
    edited_df = st.data_editor(
        rank_df,
        width='stretch',
        num_rows="dynamic",  # Allow adding/deleting rows
        height=500,
        column_config={
            "Spieler": st.column_config.TextColumn("Spieler", required=True),
            "Rang": st.column_config.NumberColumn(
                "Rang",
                min_value=1,
                max_value=6,
                required=True,
                help="1 = Stärkster, 6 = Schwächster"
            )
        },
        hide_index=True,
        key="rank_editor"
    )

    # Show rank distribution
    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("📈 Verteilung nach Rang")
        rank_counts = edited_df["Rang"].value_counts().sort_index()
        rank_dist = pd.DataFrame({
            "Rang": range(1, 7),
            "Anzahl": [rank_counts.get(i, 0) for i in range(1, 7)]
        })
        st.dataframe(rank_dist, width='stretch', hide_index=True)

    with col2:
        st.subheader("✅ Statistiken")
        st.metric("Spieler mit Rang", len(edited_df))
        st.metric("Spieler ohne Rang", len(missing_players))
        st.metric("Gesamt Spieler", len(all_players_from_prefs))

    # Save buttons
    st.markdown("---")
    col_save, col_download, col_reset = st.columns(3)

    with col_save:
        if st.button("💾 Rankings speichern", width='stretch', type="primary"):
            try:
                # Update session state
                new_ranks = {}
                for _, row in edited_df.iterrows():
                    player = str(row["Spieler"]).strip()
                    rank = int(row["Rang"])
                    if player:  # Ignore empty rows
                        new_ranks[player] = rank

                st.session_state.edited_ranks = new_ranks

                # Save to CSV
                rank_save_df = pd.DataFrame([
                    {"Spieler": player, "Rank": rank}
                    for player, rank in sorted(new_ranks.items())
                ])
                csv_bytes = rank_save_df.to_csv(index=False).encode("utf-8")

                # Save to GitHub
                msg = f"Update player rankings (edited in Rankings tab)\n\n🎾 {len(new_ranks)} players ranked"
                github_put_file(csv_bytes, msg, RANK_FILE)

                st.success(f"✅ Rankings erfolgreich gespeichert! {len(new_ranks)} Spieler aktualisiert.")
                st.info("🔄 Lade die Seite neu, um die aktualisierten Rankings zu verwenden.")
                st.balloons()
            except Exception as e:
                st.error(f"Fehler beim Speichern: {e}")

    with col_download:
        # Download CSV
        rank_save_df = pd.DataFrame([
            {"Spieler": player, "Rank": int(rank)}
            for _, (player, rank) in edited_df.iterrows()
            if str(player).strip()
        ])
        csv_bytes = rank_save_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "📥 Rankings als CSV",
            data=csv_bytes,
            file_name=RANK_FILE,
            mime="text/csv",
            width='stretch'
        )

    with col_reset:
        if st.button("🔄 Änderungen verwerfen", width='stretch'):
            st.session_state.edited_ranks = RANK.copy()
            st.rerun()

    # Help section
    st.markdown("---")
    with st.expander("ℹ️ Hilfe: Wie verwende ich die Rankings-Verwaltung?"):
        st.markdown("""
        ### Rankings bearbeiten:
        1. **Rang ändern:** Doppelklick auf eine Rang-Zelle → neue Zahl eingeben (1-6)
        2. **Spieler hinzufügen:** Klicke auf die leere Zeile am Ende der Tabelle
        3. **Spieler löschen:** Wähle Zeile aus und drücke Entf-Taste

        ### Rang-Bedeutung:
        - **Rang 1:** Stärkster Spieler (z.B. Patrick Buehrsch, Bjoern Junker)
        - **Rang 2-3:** Starke Spieler
        - **Rang 4-5:** Durchschnittliche Spieler
        - **Rang 6:** Anfänger/Schwächste Spieler

        ### Regeln:
        - **Einzel:** Rang-Differenz max 2 (z.B. Rang 2 vs Rang 4 ✅)
        - **Doppel:** Rang-Spread max 3 (z.B. Team mit Rang 1,2,3,4 ✅)

        ### Beispiel: Spieler verbessert sich:
        - Tobias Kahl spielt besser → Rang von 5 auf 4 ändern
        - Speichern → Rankings aktualisiert!
        """)

with tab_player:
    st.header("👤 Spieler-Profil")
    st.caption("Detaillierte Informationen und geplante Matches für jeden Spieler")

    # Player selection dropdown
    if all_players:
        selected_player = st.selectbox(
            "Wähle einen Spieler:",
            options=sorted(all_players),
            index=0
        )

        if selected_player:
            st.markdown("---")

            # Get player info from preferences
            player_prefs = df_prefs[df_prefs["Spieler"] == selected_player]

            if not player_prefs.empty:
                player_row = player_prefs.iloc[0]

                # Display player information in columns
                col1, col2 = st.columns(2)

                with col1:
                    st.subheader("📋 Spieler-Information")

                    # Rank
                    player_rank = RANK.get(selected_player, "Nicht festgelegt")
                    st.metric("Rang", player_rank, help="1 = Stärkster, 6 = Schwächster")

                    # Preference
                    pref = player_row.get("Preference", "keine Präferenz")
                    pref_emoji = "🎾" if pref == "nur Einzel" else "👥" if pref == "nur Doppel" else "🔄"
                    st.write(f"**Spielart-Präferenz:** {pref_emoji} {pref}")

                    # Available days
                    avail_days = available_days.get(selected_player, set())
                    if avail_days:
                        days_str = ", ".join(sorted(avail_days))
                        st.write(f"**Verfügbare Tage:** {days_str}")
                    else:
                        st.write("**Verfügbare Tage:** Keine Angabe")

                with col2:
                    st.subheader("📅 Urlaub & Notizen")

                    # Holidays
                    player_holidays = holidays.get(selected_player, [])
                    if player_holidays:
                        st.write("**Urlaub/Gesperrt:**")
                        for start, end in player_holidays:
                            if start == end:
                                st.write(f"  • {start.strftime('%d.%m.%Y')}")
                            else:
                                st.write(f"  • {start.strftime('%d.%m.%Y')} → {end.strftime('%d.%m.%Y')}")
                    else:
                        st.write("**Urlaub:** Keine Einträge")

                    # Notes
                    notes = player_row.get("Notes", "")
                    if notes and str(notes).strip() and str(notes) != "nan":
                        st.write("**Notizen:**")
                        st.info(str(notes))

                # Get player's matches from current plan
                st.markdown("---")
                st.subheader("🎾 Geplante Matches")

                # Use the working plan or result plan
                display_plan = st.session_state.get("df_result", st.session_state.df_work)

                if len(display_plan) > 0:
                    # Filter matches with this player
                    player_matches = display_plan[
                        display_plan["Spieler"].str.contains(fr"\b{re.escape(selected_player)}\b", regex=True)
                    ].copy()

                    if len(player_matches) > 0:
                        st.write(f"**{len(player_matches)} Matches geplant**")

                        # Show statistics
                        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)

                        with col_stat1:
                            einzel_count = player_matches[player_matches["Typ"].str.contains("Einzel", case=False)].shape[0]
                            st.metric("Einzel", einzel_count)

                        with col_stat2:
                            doppel_count = player_matches[player_matches["Typ"].str.contains("Doppel", case=False)].shape[0]
                            st.metric("Doppel", doppel_count)

                        with col_stat3:
                            # Count by month
                            if "Datum_dt" in player_matches.columns:
                                unique_months = player_matches["Datum_dt"].dt.to_period('M').nunique()
                                st.metric("Monate aktiv", unique_months)

                        with col_stat4:
                            # Avg per week
                            if "Woche" in player_matches.columns:
                                unique_weeks = player_matches["Woche"].nunique()
                                if unique_weeks > 0:
                                    avg_per_week = len(player_matches) / unique_weeks
                                    st.metric("Ø pro Woche", f"{avg_per_week:.1f}")

                        # Show ratio for "keine Präferenz" players
                        pref = player_row.get("Preference", "keine Präferenz")
                        if pref == "keine Präferenz" and len(player_matches) > 0:
                            singles_ratio = (einzel_count / len(player_matches)) * 100
                            doubles_ratio = (doppel_count / len(player_matches)) * 100

                            st.markdown("---")
                            st.write("**Match-Typ Verteilung** (Ziel: 35% Einzel / 65% Doppel)")

                            col_r1, col_r2 = st.columns(2)
                            with col_r1:
                                ratio_status = "✅" if 30 <= singles_ratio <= 40 else "⚠️"
                                st.metric(f"Einzel {ratio_status}", f"{singles_ratio:.1f}%")
                            with col_r2:
                                ratio_status = "✅" if 60 <= doubles_ratio <= 70 else "⚠️"
                                st.metric(f"Doppel {ratio_status}", f"{doubles_ratio:.1f}%")

                        st.markdown("---")

                        # Show matches table
                        display_matches = player_matches[["Datum", "Tag", "Slot", "Typ", "Spieler"]].copy()

                        # Add teammates/opponents column
                        def get_partners(row):
                            all_players = [p.strip() for p in str(row["Spieler"]).split(",")]
                            partners = [p for p in all_players if p != selected_player]
                            return ", ".join(partners)

                        display_matches["Mitspieler"] = display_matches.apply(get_partners, axis=1)

                        # Reorder columns
                        display_matches = display_matches[["Datum", "Tag", "Slot", "Typ", "Mitspieler"]]

                        # Display table
                        st.dataframe(
                            display_matches,
                            width='stretch',
                            height=400,
                            hide_index=True
                        )

                        # Download option
                        csv_bytes = display_matches.to_csv(index=False).encode("utf-8")
                        st.download_button(
                            f"📥 {selected_player}'s Matches als CSV",
                            data=csv_bytes,
                            file_name=f"{selected_player.replace(' ', '_')}_Matches_2026.csv",
                            mime="text/csv"
                        )
                    else:
                        st.info(f"🔍 Keine Matches für {selected_player} im aktuellen Plan gefunden.")
                else:
                    st.info("📭 Der Plan ist noch leer. Verwende die Auto-Population, um Slots zu füllen.")
            else:
                st.warning(f"⚠️ Keine Präferenzen für {selected_player} gefunden in {PREFS_FILE}")
    else:
        st.warning("⚠️ Keine Spieler gefunden. Bitte lade die Preferences-Datei.")

# Documentation
st.markdown("---")
with st.expander("ℹ️ Saison-Informationen 2026"):
    st.markdown(f"""
    ### Winter-Saison 2026

    **Zeitraum:** {SEASON_START.strftime('%d.%m.%Y')} - {SEASON_END.strftime('%d.%m.%Y')}

    **Trainingsdateien:**
    - Plan: `{PLAN_FILE}`
    - Präferenzen: `{PREFS_FILE}`
    - Rankings: `{RANK_FILE}` (1=stärkster, 6=schwächster)

    **Trainingstage:**
    - **Montag:** 2x Doppel (20:00-22:00)
    - **Mittwoch:** 3x Einzel + 2x Doppel (18:00-21:30)
    - **Donnerstag:** 2x Einzel (20:00-21:30)

    **Gesamt pro Woche:** 7 Slots = 14 Einzel-Plätze + 12 Doppel-Plätze = 26 Spieler-Plätze/Woche

    **Saison-Statistik:**
    - Wochen: ca. 17 Wochen (Januar-April)
    - Theoretisch möglich: ~119 Slots
    - Spieler-Plätze gesamt: ~442

    **Berücksichtigt:**
    - Urlaube aus CSV (2026-01-01 bis 2026-04-26)
    - Spielerpräferenzen (nur Einzel/nur Doppel)
    - Verfügbarkeit nach Wochentagen
    - Spieler-Rankings (Einzel: Rang-Differenz ≤ 2, Doppel: Rang-Differenz ≤ 3)
    - Load Balancing für faire Verteilung
    """)

with st.expander("🏆 Spieler-Rankings"):
    st.markdown("""
    **Ranking-System:** 1 (stärkster) bis 6 (schwächster)

    **Regel für Einzel:** Rang-Differenz zwischen 2 Spielern ≤ 2

    **Regel für Doppel:** Rang-Differenz zwischen stärkster und schwächster Spieler ≤ 3

    **Quelle:** `Player_Ranks_2026.csv` (aus audit prompt.txt)
    """)

    # Show rank distribution
    rank_counts = {}
    for player, rank in RANK.items():
        rank_counts[rank] = rank_counts.get(rank, 0) + 1

    rank_df = pd.DataFrame([
        {"Rang": 1, "Beschreibung": "Stärkster", "Anzahl Spieler": rank_counts.get(1, 0)},
        {"Rang": 2, "Beschreibung": "Sehr Stark", "Anzahl Spieler": rank_counts.get(2, 0)},
        {"Rang": 3, "Beschreibung": "Stark", "Anzahl Spieler": rank_counts.get(3, 0)},
        {"Rang": 4, "Beschreibung": "Überdurchschnittlich", "Anzahl Spieler": rank_counts.get(4, 0)},
        {"Rang": 5, "Beschreibung": "Durchschnittlich", "Anzahl Spieler": rank_counts.get(5, 0)},
        {"Rang": 6, "Beschreibung": "Schwächster", "Anzahl Spieler": rank_counts.get(6, 0)},
    ])
    st.dataframe(rank_df, width='stretch', hide_index=True)
